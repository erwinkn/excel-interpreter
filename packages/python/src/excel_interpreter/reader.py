from typing import Any, Callable, Literal, cast

import pandas as pd
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.worksheet.worksheet import Worksheet
from rapidfuzz import fuzz, process
from excel_interpreter.interpreter import ExcelInterpreter
from excel_interpreter.utils import column_as_int
from openpyxl.worksheet.formula import ArrayFormula



def remove_excel_illegal_characters(df):
    for col in df.columns:
        if df[col].dtype == object:  # Only process string columns
            df[col] = df[col].str.replace(ILLEGAL_CHARACTERS_RE, "", regex=True)

    return df


class ExcelReader:
    def __init__(self, ws: Worksheet, row=1, col=1, auto_evaluate=False) -> None:
        self.ws = ws
        self.row = row
        self.col = col
        self.max_row = ws.max_row
        self.max_column = ws.max_column
        if auto_evaluate:
            assert ws.parent is not None
            self.interpreter = ExcelInterpreter(ws.parent)
        else:
            self.interpreter = None

        self.labels: dict[str, list[tuple[int, int]]] = {}
        for row in range(1, self.max_row + 1):
            for col in range(1, self.max_column + 1):
                value = self.ws.cell(row, col).value
                if not isinstance(value, str):
                    continue
                # Skip formulas
                if value.startswith("="):
                    continue
                if len(value) < 3:
                    continue
                if value not in self.labels:
                    self.labels[value] = []
                self.labels[value].append((row, col))
        # For performance
        self.label_keys = list(self.labels.keys())

    def find_cell(
        self,
        label: str,
        similarity: float = 0.9,
        row: int | None = None,
        col: int | None = None,
        result_idx: int = 0
    ) -> tuple[int, int]:
        coords = self.find_cell_optional(label, similarity=similarity, row=row, col=col, result_idx=result_idx)
        if coords is None:
            raise ValueError(f"Label {label} not found (similarity = {similarity})")
        return coords

    def find_cell_optional(
        self,
        label: str,
        similarity: float = 0.9,
        row: int | None = None,
        col: int | None = None,
        safe=False,
        ignore_spaces: bool = False,
        result_idx=0
    ) -> tuple[int, int] | None:
        cells = []
        if ignore_spaces:
            stripped = {}
            for key, value in self.labels.items():
                new_key = key.replace(" ", "")
                if new_key in stripped:
                    stripped[new_key].extend(value)
                else:
                    stripped[new_key] = value
            self.labels = stripped
        if label in self.labels:
            cells = self.labels[label]
        else:
            matches = process.extract(
                label, self.label_keys, scorer=fuzz.ratio, limit=1
            )
            matches = list(sorted(matches, key=lambda m: m[1]))
            for match in matches:
                if match[1] < similarity * 100:
                    continue
                cells.extend(self.labels[match[0]])

        if row is not None:
            cells = [cell for cell in cells if cell[0] == row]
        if col is not None:
            cells = [cell for cell in cells if cell[1] == col]
        if len(cells) == 0:
            return None
        # if len(cells) > 1:
        #     if not safe:
        #         raise ValueError(
        #             f'More than one cell matched the query "{label}": {cells}'
        #         )
        #     else:
        #         return None
        return cells[result_idx]

    def read(
        self,
        label: str,
        position: Literal["top", "bottom", "left", "right"],
        conversion: Callable | None = None,
        similarity=0.7,
        row: int | None = None,
        col: int | None = None,
        skip: int = 0,
    ):
        value = self.read_optional(
            label,
            position,
            conversion=conversion,
            similarity=similarity,
            row=row,
            col=col,
            skip=skip,
        )
        if value is None:
            raise ValueError(
                f'Value for "{label}" not found (position={position}, similarity={similarity})'
            )
        return value

    def read_at(
        self,
        row: int,
        col: int,
        conversion: Callable | None = None,
        safe: bool = False,
        link: bool = False,
    ) -> Any:
        cell = self.ws.cell(row, col)
        if link:
            value = cell.hyperlink.target if cell.hyperlink else None
        else:
            value = cell.value

        if value == "":
            value = None

        if (
            self.interpreter
            and isinstance(value, str)
            and value.strip().startswith("=")
        ):
            try:
                value = self.interpreter.evaluate(value, current_sheet=self.ws)
                if value is not None and conversion:
                    value = conversion(value)
            except Exception as e:
                if safe:
                    # print(f"Error evaluating {cell.coordinate}: {value}: {e}")
                    value = pd.NA
                else:
                    raise
        elif isinstance(value, ArrayFormula):
            try:
                value = self.interpreter.evaluate(value.text, current_sheet=self.ws)[0]
                if value is not None and conversion:
                    value = conversion(value)
            except TypeError:
                try:
                    value = self.interpreter.evaluate(value.text, current_sheet=self.ws)
                except:
                    pass
            except Exception as e:
                if safe:
                    value = pd.NA
                else:
                    raise

        return value

    def read_optional(
        self,
        label: str,
        position: Literal["top", "bottom", "left", "right"],
        conversion: Callable | None = None,
        safe=False,
        label_optional=False,
        similarity=0.7,
        link=False,
        row: int | None = None,
        col: int | None = None,
        skip: int = 0,  # number of cells to skip between label and value
    ):
        if label_optional:
            coords = self.find_cell_optional(
                label, similarity=similarity, row=row, col=col, safe=safe
            )
            if coords is None:
                return None
        else:
            coords = self.find_cell(label, similarity=similarity)

        row, col = coords
        if position == "top":
            target_row, target_col = row - skip - 1, col
        elif position == "bottom":
            target_row, target_col = row + skip + 1, col
        elif position == "left":
            target_row, target_col = row, col - skip - 1
        elif position == "right":
            target_row, target_col = row, col + skip + 1
        else:
            raise ValueError(f"Invalid position: {position}")

        return self.read_at(target_row, target_col, conversion, safe, link)

    def read_str(
        self,
        label: str,
        position: Literal["top", "bottom", "left", "right"],
        similarity=0.7,
        row: int | None = None,
        col: int | None = None,
    ) -> str:
        return cast(str, self.read(label, position, str, similarity, row=row, col=col))

    def read_str_optional(
        self,
        label: str,
        position: Literal["top", "bottom", "left", "right"],
        label_optional=False,
        safe=False,
        similarity=0.7,
        row: int | None = None,
        col: int | None = None,
        skip: int = 0,
    ) -> str | None:
        return cast(
            str | None,
            self.read_optional(
                label,
                position,
                str,
                label_optional=label_optional,
                safe=safe,
                similarity=similarity,
                row=row,
                col=col,
                skip=skip,
            ),
        )

    def read_float(
        self,
        label: str,
        position: Literal["top", "bottom", "left", "right"],
        similarity=0.7,
        row: int | None = None,
        col: int | None = None,
    ) -> float:
        return cast(
            float,
            self.read(label, position, float, similarity=similarity, row=row, col=col),
        )

    def read_float_optional(
        self,
        label: str,
        position: Literal["top", "bottom", "left", "right"],
        label_optional=False,
        safe=False,
        similarity=0.7,
        row: int | None = None,
        col: int | None = None,
        skip: int = 0,
    ) -> float | None:
        return cast(
            float | None,
            self.read_optional(
                label,
                position,
                float,
                safe=safe,
                label_optional=label_optional,
                similarity=similarity,
                row=row,
                col=col,
                skip=skip,
            ),
        )

    def read_int(
        self,
        label: str,
        position: Literal["top", "bottom", "left", "right"],
        similarity=0.7,
        row: int | None = None,
        col: int | None = None,
    ) -> int:
        return cast(
            int,
            self.read(label, position, int, similarity=similarity, row=row, col=col),
        )

    def read_int_optional(
        self,
        label: str,
        position: Literal["top", "bottom", "left", "right"],
        label_optional=False,
        safe=False,
        similarity=0.7,
        row: int | None = None,
        col: int | None = None,
        skip: int = 0,
    ) -> int | None:
        return cast(
            int | None,
            self.read_optional(
                label,
                position,
                int,
                label_optional=label_optional,
                safe=safe,
                similarity=similarity,
                row=row,
                col=col,
                skip=skip,
            ),
        )

    def read_datetime(
        self,
        label: str,
        position: Literal["top", "bottom", "left", "right"],
        similarity=0.7,
        row: int | None = None,
        col: int | None = None,
    ) -> pd.Timestamp:
        return cast(
            pd.Timestamp,
            self.read(
                label, position, pd.to_datetime, similarity=similarity, row=row, col=col
            ),
        )

    def read_datetime_optional(
        self,
        label: str,
        position: Literal["top", "bottom", "left", "right"],
        label_optional=False,
        safe=False,
        similarity=0.7,
        row: int | None = None,
        col: int | None = None,
        skip: int = 0,
    ) -> pd.Timestamp | None:
        return cast(
            pd.Timestamp | None,
            self.read_optional(
                label,
                position,
                pd.to_datetime,
                label_optional=label_optional,
                safe=safe,
                similarity=similarity,
                row=row,
                col=col,
                skip=skip,
            ),
        )

    def determine_table_extent(
        self,
        row: int,
        col: int | str,
        end_row_label="",
        end_col_label="",
    ):
        # Find the end row using end_row_label
        start_col = end_col = column_as_int(col)
        start_row = end_row = row

        for end_col in range(start_col, self.max_column + 1):
            value = self.ws.cell(start_row, end_col).value
            if end_col_label == value:
                break
            # We consider the empty string to be the same thing as None
            if end_col_label == "" and (value is None or value == ""):
                break
        else:
            # Loop completed without breaking - include the last column
            end_col += 1

        for end_row in range(start_row, self.max_row + 1):
            value = self.ws.cell(end_row, start_col).value
            if end_row_label == value:
                break
            if end_row_label == "" and (value is None or value == ""):
                break
        else:
            # Loop completed without breaking - include the last row
            end_row += 1

        return start_row, end_row, start_col, end_col

    def read_table(
        self,
        row: int,
        col: int | str,
        end_row_label="",
        end_col_label="",
        end_row_index: int | None = None,
        end_col_index: int | None = None,
        safe=False,
    ) -> list[list]:
        start_row, end_row, start_col, end_col = self.determine_table_extent(
            row, col, end_row_label=end_row_label, end_col_label=end_col_label
        )
        if end_row_index:
            end_row = end_row_index
        if end_col_index:
            end_col = end_col_index

        # Iterate over the table and collect data
        data = []
        for row in range(start_row, end_row):
            row_data = []
            for col in range(start_col, end_col):
                row_data.append(self.read_at(row, col, safe=safe))
            data.append(row_data)
        return data

    def read_df(
        self,
        row: int,
        col: int | str,
        end_row_label="",
        end_col_label="",
        end_row_index: int | None = None,
        end_col_index: int | None = None,
        header=False,
        columns: list[str] | None = None,
        safe=False
    ) -> pd.DataFrame:
        # Iterate over the table and collect data
        data = self.read_table(
            row,
            col,
            end_row_label=end_row_label,
            end_col_label=end_col_label,
            end_row_index=end_row_index,
            end_col_index=end_col_index,
            safe=safe,
        )

        if header:
            columns = data[0]
            data = data[1:]

        return pd.DataFrame(data, columns=columns)
    
