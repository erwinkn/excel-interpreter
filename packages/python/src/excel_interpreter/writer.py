from typing import Literal, Sequence, overload
from copy import copy
import typing
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Alignment,
    Protection,
    Side,
)
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import column_index_from_string, get_column_letter

from .expression import Constant, Expression, ExprOrValue


RowLike = Sequence[ExprOrValue | None]
ColumnLike = Sequence[ExprOrValue | None]

BorderStyle = Literal[
    "dashDot",
    "dashDotDot",
    "dashed",
    "dotted",
    "double",
    "hair",
    "medium",
    "mediumDashDot",
    "mediumDashDotDot",
    "mediumDashed",
    "slantDashDot",
    "thick",
    "thin",
]
border_styles: Sequence[str] = typing.get_args(BorderStyle)


class Row:
    def __init__(
        self,
        *cells: ExprOrValue | None,
        # All the styling options available for individual cells in `openpyxl`
        # https://openpyxl.readthedocs.io/en/stable/styles.html
        number_format: str | None = None,
        font: Font | None = None,
        fill: PatternFill | None = None,
        border: Border | None = None,
        alignment: Alignment | None = None,
        protection: Protection | None = None,
    ) -> None:
        self.cells = [
            Constant(cell) if isinstance(cell, (int, float, str, bool)) else cell
            for cell in cells
        ]
        self.number_format = number_format
        self.font = font
        self.fill = fill
        self.border = border
        self.alignment = alignment
        self.protection = protection

    def __len__(self):
        return len(self.cells)

    def __getitem__(self, key):
        if isinstance(key, int):
            return self.cells[key]
        else:
            raise NotImplementedError()


# Identical to Row
class Column:
    def __init__(
        self,
        *cells: ExprOrValue | None,
        # All the styling options available for individual cells in `openpyxl`
        # https://openpyxl.readthedocs.io/en/stable/styles.html
        number_format: str | None = None,
        font: Font | None = None,
        fill: PatternFill | None = None,
        border: Border | None = None,
        alignment: Alignment | None = None,
        protection: Protection | None = None,
    ) -> None:
        self.cells = cells
        self.number_format = number_format
        self.font = font
        self.fill = fill
        self.border = border
        self.alignment = alignment
        self.protection = protection

    def __len__(self):
        return len(self.cells)

    def __getitem__(self, key):
        if isinstance(key, int):
            return self.cells[key]
        else:
            raise NotImplementedError()


class ExcelWriter:
    ws: Worksheet
    row: int
    col: int

    def __init__(
        self,
        ws: Worksheet,
        row: int = 1,
        col: int | str = 1,
    ) -> None:
        self.ws = ws
        self.row = row
        if isinstance(col, str):
            col = column_index_from_string(col)
        self.col = col

    def move(self, row: int | None = None, col: int | str | None = None):
        if row is not None:
            self.row = row
        if col is not None:
            if isinstance(col, str):
                col = column_index_from_string(col)
            self.col = col

    def draw_top_border(
        self,
        *,
        length: int,
        start: int | str | None = None,
        row: int | None = None,
        style: BorderStyle = "thin",
        color: str = "000000",
    ):
        if start is None:
            start = self.col
        if row is None:
            row = self.row
        if isinstance(start, str):
            start = column_index_from_string(start)
        end = start + length

        for col in range(start, end):
            cell = self.ws.cell(row, col)
            current_border: Border | None = cell.border
            if current_border is not None:
                border = copy(current_border)
                border.top = Side(style=style, color=color)
                cell.border = border
            else:
                cell.border = Border(top=Side(style=style, color=color))

    def draw_bottom_border(
        self,
        *,
        length: int,
        start: int | str | None = None,
        row: int | None = None,
        style: BorderStyle = "thin",
        color: str = "000000",
    ):
        if start is None:
            start = self.col
        if row is None:
            row = self.row
        if isinstance(start, str):
            start = column_index_from_string(start)
        end = start + length

        for col in range(start, end + 1):
            cell = self.ws.cell(row, col)
            current_border: Border | None = cell.border
            if current_border is not None:
                border = copy(current_border)
                border.bottom = Side(style=style, color=color)
                cell.border = border
            else:
                cell.border = Border(bottom=Side(style=style, color=color))

    def write_row(
        self,
        row: RowLike | Row,
        *,
        # All the styling options available for individual cells in `openpyxl`
        # https://openpyxl.readthedocs.io/en/stable/styles.html
        number_format: str | None = None,
        font: Font | None = None,
        fill: PatternFill | None = None,
        border: Border | None = None,
        alignment: Alignment | None = None,
        protection: Protection | None = None,
    ):
        """Write a row to the spreadsheet.

        Optionally, set default styling for the row. Cell objects with their own styling will have priority over the default row styling
        """
        if isinstance(row, Row):
            number_format = row.number_format or number_format
            font = row.font or font
            fill = row.fill or fill
            border = row.border or border
            alignment = row.alignment or alignment
            protection = row.protection or protection
            row = row.cells

        return_row = []
        for i, cell in enumerate(row):
            if cell is None:
                continue
            if isinstance(cell, Expression):
                return_row.append(
                    cell.write(
                        self.ws,
                        row=self.row,
                        col=self.col + i,
                        number_format=number_format,
                        font=font,
                        fill=fill,
                        border=border,
                        alignment=alignment,
                        protection=protection,
                    )
                )
            else:
                # The `value` argument isn't happy with non-string types, even though it's perfectly fine *shrug*
                excel_cell = self.ws.cell(self.row, self.col + i, value=cell)  # type: ignore
                return_row.append(excel_cell)
                # Can't set options with `None`, otherwise it will trigger an error
                # when writing the spreadsheet to a file.
                if number_format is not None:
                    excel_cell.number_format = number_format
                if font is not None:
                    excel_cell.font = font
                if fill is not None:
                    excel_cell.fill = fill
                if border is not None:
                    excel_cell.border = border
                if alignment is not None:
                    excel_cell.alignment = alignment
                if protection is not None:
                    excel_cell.protection = protection
        self.row += 1
        return return_row

    @overload
    def write_column(
        self,
        col: ColumnLike,
        *,
        # All the styling options available for individual cells in `openpyxl`
        # https://openpyxl.readthedocs.io/en/stable/styles.html
        number_format: str | None = None,
        font: Font | None = None,
        fill: PatternFill | None = None,
        border: Border | None = None,
        alignment: Alignment | None = None,
        protection: Protection | None = None,
    ): ...

    @overload
    def write_column(
        self,
        col: Column,
    ): ...

    def write_column(
        self,
        col: ColumnLike | Column,
        *,
        # All the styling options available for individual cells in `openpyxl`
        # https://openpyxl.readthedocs.io/en/stable/styles.html
        number_format: str | None = None,
        font: Font | None = None,
        fill: PatternFill | None = None,
        border: Border | None = None,
        alignment: Alignment | None = None,
        protection: Protection | None = None,
    ):
        """Write a row to the spreadsheet.

        Optionally, set default styling for the row. Cell objects with their own styling will have priority over the default row styling
        """
        if isinstance(col, Column):
            number_format = col.number_format
            font = col.font
            border = col.border
            alignment = col.alignment
            protection = col.protection
            col = col.cells

        for i, cell in enumerate(col):
            if cell is None:
                continue
            if isinstance(cell, Expression):
                cell.write(
                    self.ws,
                    row=self.row + i,
                    col=self.col,
                    number_format=number_format,
                    font=font,
                    fill=fill,
                    border=border,
                    alignment=alignment,
                    protection=protection,
                )
            else:
                # The `value` argument isn't happy with non-string types, even though it's perfectly fine *shrug*
                excel_cell = self.ws.cell(self.row + i, self.col, value=cell)  # type: ignore
                # Can't set options with `None`, otherwise it will trigger an error
                # when writing the spreadsheet to a file.
                if number_format is not None:
                    excel_cell.number_format = number_format
                if font is not None:
                    excel_cell.font = font
                if fill is not None:
                    excel_cell.fill = fill
                if border is not None:
                    excel_cell.border = border
                if alignment is not None:
                    excel_cell.alignment = alignment
                if protection is not None:
                    excel_cell.protection = protection
        self.col += 1

    def write_rows(
        self,
        *rows: Row | RowLike,
        # All the styling options available for individual cells in `openpyxl`
        # https://openpyxl.readthedocs.io/en/stable/styles.html
        number_format: str | None = None,
        font: Font | None = None,
        fill: PatternFill | None = None,
        border: Border | None = None,
        alignment: Alignment | None = None,
        protection: Protection | None = None,
    ):
        for row in rows:
            self.write_row(
                row,
                number_format=number_format,
                font=font,
                fill=fill,
                border=border,
                alignment=alignment,
                protection=protection,
            )

    def write_columns(
        self,
        *columns: Column | ColumnLike,
    ):
        for col in columns:
            self.write_column(col)

    @overload
    def write_table(
        self,
        *args: Row | RowLike,
        header_rows=1,
        label_columns=1,
        by: Literal["row"] = "row",
    ): ...

    @overload
    def write_table(
        self,
        *args: Column | ColumnLike,
        header_rows=1,
        label_columns=1,
        by: Literal["column"],
    ): ...

    def write_table(
        self,
        *args: Column | ColumnLike | Row | RowLike,
        header_rows=1,
        label_columns=1,
        by: Literal["row", "column"] = "row",
    ):
        start_row = self.row
        start_col = self.col

        if by == "row":
            nb_rows = len(args)
            nb_columns = 0
            for row in args:
                nb_columns = max(nb_columns, len(row))
                self.write_row(row)  # type: ignore
        else:
            nb_rows = 0
            nb_columns = len(args)
            for column in args:
                nb_rows = max(nb_rows, len(column))
                self.write_column(column)  # type: ignore

        for i in range(header_rows):
            for j in range(nb_columns):
                cell = self.ws.cell(start_row + i, start_col + j)
                cell.font = Font(bold=True)

        for i in range(header_rows, len(args)):
            for j in range(label_columns):
                cell = self.ws.cell(start_row + i, start_col + j)
                cell.font = Font(bold=True)

        self.row = start_row + nb_rows
        self.col = start_col

    def set_column_width(self, column: int | str, width: float):
        if isinstance(column, int):
            column = get_column_letter(column)
        self.ws.column_dimensions[column].width = width

    # def set_row_height(self, row: int, height: float):
    #     self.ws.row_dimensions[row].height = height
