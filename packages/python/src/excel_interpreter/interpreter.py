from datetime import date, datetime
from decimal import Decimal
from typing import Any, Dict, List, Optional, Tuple, Union

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText
from openpyxl.utils import (
    column_index_from_string,
    get_column_letter,
    range_boundaries,
    coordinate_to_tuple,
)
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
from openpyxl.worksheet.worksheet import Worksheet

from .ast import (
    ArrayLiteral,
    ASTNode,
    BinaryOperation,
    CellRange,
    CellReference,
    Constant,
    ExcelFunction,
    NameReference,
    UnaryOperation,
)
from excel_interpreter.errors import CycleError, NameNotFound, SheetNotFound
from excel_interpreter.functions import EXCEL_FUNCTIONS, coerce_to_number
from excel_interpreter.operators import (
    add,
    concatenate,
    divide,
    eq,
    gt,
    gte,
    lt,
    lte,
    multiply,
    neq,
    power,
    subtract,
)
from excel_interpreter.parser import ExcelParser
from excel_interpreter.tokenizer import ExcelTokenizer
from excel_interpreter.types import (
    ExcelValue,
    ScalarExcelValue,
)
from excel_interpreter.utils import array_shape, column_as_int, column_as_str

# TODO:
# - Cache evaluation results for cells rather than raw values
# - Handle ranges over array formulas
# - In general, be able to represent the value of individual cells within an array formula
# - Propagating errors (important for array formulas where certain cells may error without others)


def check_is_scalar(x: ExcelValue) -> ScalarExcelValue:
    if isinstance(x, list):
        raise NotImplementedError(
            "You hit a case where the interpreter is not able to handle array formulas yet. "
            "Please reach out to Erwin."
        )
    return x


def _format_ref(node: ASTNode, sheet: Optional[Union[str, Worksheet]] = None) -> str:
    """Format a node for error messages, optionally including sheet context."""
    if isinstance(node, CellReference):
        # Use the node's sheet if available, otherwise use the provided sheet
        node_sheet = node.sheet
        if not node_sheet and sheet:
            node_sheet = sheet.title if isinstance(sheet, Worksheet) else sheet
        sheet_prefix = f"{node_sheet}!" if node_sheet else ""
        return f"{sheet_prefix}{column_as_str(node.column)}{node.row}"
    if isinstance(node, NameReference):
        sheet_info = ""
        if sheet:
            sheet_name = sheet.title if isinstance(sheet, Worksheet) else sheet
            sheet_info = f" (in {sheet_name})"
        return f"name:{node.name}{sheet_info}"
    return str(node)


class EvaluationStack:
    """Tracks the evaluation stack with both nodes and their context (sheet)."""

    def __init__(self):
        self.stack: List[Tuple[ASTNode, Union[str, Worksheet]]] = []

    def push(self, node: ASTNode, sheet: Union[str, Worksheet]) -> None:
        """Add a node and its context to the stack."""
        self.stack.append((node, sheet))

    def pop(self) -> None:
        """Remove the last node from the stack."""
        self.stack.pop()

    def contains(self, node: ASTNode, current_sheet: Union[str, Worksheet]) -> bool:
        """
        Check if the node+sheet combination is already in the stack.

        For CellReference nodes, we need to compare both the node and sheet to avoid
        false cycle detection when the same cell reference appears in different sheets.

        For other node types like NameReference, we only check the node itself since
        names should be unique across the workbook.
        """
        if isinstance(node, CellReference):
            # For cell references, check both node and sheet
            # Normalize sheet names for comparison
            current_sheet_name = (
                current_sheet.title
                if isinstance(current_sheet, Worksheet)
                else current_sheet
            )

            for stack_node, stack_sheet in self.stack:
                if not isinstance(stack_node, CellReference):
                    continue

                # If the node has an explicit sheet, use that for comparison
                if node.sheet:
                    node_matches = node == stack_node and node.sheet == stack_node.sheet
                else:
                    # If no explicit sheet, compare with the context sheet
                    stack_sheet_name = (
                        stack_sheet.title
                        if isinstance(stack_sheet, Worksheet)
                        else stack_sheet
                    )
                    node_matches = (
                        node == stack_node and current_sheet_name == stack_sheet_name
                    )

                if node_matches:
                    return True
            return False
        else:
            # For other node types, just check the node itself
            return any(n == node for n, _ in self.stack)

    def format_cycle_path(self, node: ASTNode) -> str:
        """Format the evaluation stack into a readable cycle path with sheet context."""
        path = [_format_ref(n, s) for n, s in self.stack]
        # For the current node that caused the cycle, we don't have its sheet in the stack yet
        # Use the sheet from the last item in the stack as an approximation
        last_sheet = self.stack[-1][1] if self.stack else None
        path.append(_format_ref(node, last_sheet))
        return " -> ".join(path)


class ExcelInterpreter:
    def __init__(self, workbook: Workbook):
        self.workbook = workbook
        self.cache: dict[tuple[str, int, int], Any] = {}
        # Worksheet -> name -> cell reference or range
        # "" is used as the key for the global scope
        self.defined_names: dict[str, dict[str, CellRange | CellReference]] = {"": {}}
        # Sheet -> affected cell (row, col) -> array formula cell (row, col)
        self.array_formulae: dict[tuple[str, int, int], tuple[int, int]] = {}
        self.load_defined_names()
        self.load_array_formulae()
        self.evaluation_stack = EvaluationStack()

    def reload(self):
        self.cache.clear()
        self.load_defined_names()
        self.load_array_formulae()

    def evaluate(
        self,
        formula_or_node: Union[str, ASTNode],
        current_sheet: str | Worksheet,
    ) -> Any:
        """Evaluate a formula or AST node in the context of the current
        sheet."""
            
        # Parse the formula once
        if isinstance(formula_or_node, str):
            if not formula_or_node.startswith("="):
                # Raw string value, not a formula
                return formula_or_node
            formula_or_node = formula_or_node.replace("_xlfn.","").replace("_xlws.","")
            if formula_or_node.startswith("=STDEV.P"):
                formula_or_node = formula_or_node.replace("STDEV.P", "STDEV_P")
            tokens = ExcelTokenizer(formula_or_node).tokenize()
            node = ExcelParser(tokens).parse()
        else:
            node = formula_or_node
        return self._evaluate_node(node, current_sheet)

    def _evaluate_node(
        self, node: ASTNode, current_sheet: str | Worksheet
    ) -> ExcelValue:
        """Evaluate an AST node in the context of the current sheet."""

        if isinstance(node, Constant):
            return node.value

        elif isinstance(node, ArrayLiteral):
            return self._evaluate_array(node, current_sheet)

        elif isinstance(node, BinaryOperation):
            return self._evaluate_binary_op(node, current_sheet)

        elif isinstance(node, UnaryOperation):
            return self._evaluate_unary_op(node, current_sheet)

        elif isinstance(node, CellReference):
            return self._evaluate_cell_ref(node, current_sheet)

        elif isinstance(node, CellRange):
            return self._evaluate_cell_range(node, current_sheet)

        elif isinstance(node, ExcelFunction):
            return self._evaluate_function(node, current_sheet)

        elif isinstance(node, NameReference):
            return self._evaluate_name_ref(node, current_sheet)

        raise ValueError(f"Unknown node type: {type(node)}")

    def _evaluate_binary_op(
        self, node: BinaryOperation, current_sheet: str | Worksheet
    ) -> ExcelValue:
        """Evaluate a binary operation."""
        left = self._evaluate_node(node.left, current_sheet)
        right = self._evaluate_node(node.right, current_sheet)

        match node.operator:
            case "+":
                return add(left, right)
            case "-":
                return subtract(left, right)
            case "*":
                return multiply(left, right)
            case "/":
                try:
                    return divide(left, right)
                except ZeroDivisionError:
                    return "#DIV/0!"
            case "&":
                return concatenate(left, right)
            case "^":
                return power(left, right)
            case "=":
                return eq(left, right)
            case "<>":
                return neq(left, right)
            case "<":
                return lt(left, right)
            case ">":
                return gt(left, right)
            case "<=":
                return lte(left, right)
            case ">=":
                return gte(left, right)
            case _:
                raise ValueError(f"Unknown operator: {node.operator}")

    def _evaluate_unary_op(
        self, node: UnaryOperation, current_sheet: str | Worksheet
    ) -> ExcelValue:
        """Evaluate a unary operation."""
        value = self._evaluate_node(node.operand, current_sheet)

        match node.operator:
            case "+":
                return value
            case "-":
                # Negate the number
                return -coerce_to_number(value)
            case _:
                raise ValueError(f"Unknown unary operator: {node.operator}")

    def _evaluate_cell_ref(
        self, node: CellReference, current_sheet: str | Worksheet
    ) -> ExcelValue:
        """Evaluate a cell reference."""
        sheet = self._get_sheet(node.sheet or current_sheet)

        cache_key = (sheet.title, node.row, column_as_int(node.column))
        cached_result = self.cache.get(cache_key)
        if cached_result is not None:
            return cached_result

        if self.evaluation_stack.contains(node, current_sheet):
            cycle_path = self.evaluation_stack.format_cycle_path(node)
            raise CycleError(f"Detected cycle: {cycle_path}")
        self.evaluation_stack.push(node, current_sheet)

        if cache_key in self.array_formulae:
            array_formula_row, array_formula_col = self.array_formulae[cache_key]
            array_formula = sheet.cell(array_formula_row, array_formula_col).value
            assert isinstance(array_formula, ArrayFormula), (
                f"Expected ArrayFormula at cell {column_as_str(array_formula_col)}{array_formula_row}"
            )
            self._apply_array_formula(array_formula, sheet)
            cached_result = self.cache.get(cache_key)
            assert cached_result is not None, (
                f"Evaluated array formula {array_formula} for cell {column_as_str(array_formula_col)}{array_formula_row}, but it produced no result."
            )
            return cached_result

        cell = sheet.cell(row=node.row, column=column_as_int(node.column))
        value = cell.value
        if isinstance(value, ArrayFormula):
            # openpyxl also gives us the range of affected cells, but I think we
            # don't care, we can evaluate array formulas without it.
            value = value.text
        if isinstance(value, Decimal):
            value = float(value)
        if isinstance(value, date) and not isinstance(value, datetime):
            value = datetime(value.year, value.month, value.day)
        if isinstance(value, DataTableFormula):
            raise NotImplementedError(
                "Support for DataTableFormula not implemented yet"
            )
        if isinstance(value, CellRichText):
            value = str(value)
        if isinstance(value, str) and value.startswith("="):
            formula = value
            value = self.evaluate(formula, current_sheet=sheet)
            # print(f"{column_as_str(node.column)}{node.row} = {value} ({formula})")

        self.cache[cache_key] = value
        self.evaluation_stack.pop()
        return value

    def _evaluate_cell_range(
        self, node: CellRange, current_sheet: str | Worksheet
    ) -> list[ExcelValue] | list[list[ExcelValue]]:
        """Evaluate a cell range, returning an array value."""
        start_col = column_as_int(node.start.column)
        end_col = column_as_int(node.end.column)
        start_row = node.start.row
        end_row = node.end.row

        # Use the sheet from the range if specified, otherwise use current sheet
        sheet_name = node.start.sheet or (
            current_sheet.title
            if isinstance(current_sheet, Worksheet)
            else current_sheet
        )

        # I'm sure there's a more elegant way to handle the single-cell + two 1D
        # cases, but I lack time and this is pretty foolproof.
        if start_row == end_row and start_col == end_col:
            # Single cell
            return [
                self._evaluate_cell_ref(
                    CellReference(start_col, start_row, sheet=sheet_name),
                    current_sheet=current_sheet,
                )
            ]

        if start_row == end_row:
            # 1D array
            return [
                check_is_scalar(
                    self._evaluate_cell_ref(
                        CellReference(column=col, row=start_row, sheet=sheet_name),
                        current_sheet=current_sheet,
                    )
                )
                for col in range(start_col, end_col + 1)
            ]
        if start_col == end_col:
            # 1D array
            return [
                check_is_scalar(
                    self._evaluate_cell_ref(
                        CellReference(column=start_col, row=row, sheet=sheet_name),
                        current_sheet=current_sheet,
                    )
                )
                for row in range(start_row, end_row + 1)
            ]

        # 2D array
        values: list[list[ScalarExcelValue]] = []
        for row in range(node.start.row, node.end.row + 1):
            row_values: list[ScalarExcelValue] = []
            for col in range(start_col, end_col + 1):
                ref = CellReference(col, row, sheet=sheet_name)
                row_values.append(
                    check_is_scalar(self._evaluate_cell_ref(ref, current_sheet))
                )
            values.append(row_values)

        return values

    def _evaluate_function(
        self, node: ExcelFunction, current_sheet: str | Worksheet
    ) -> ExcelValue:
        """Evaluate an Excel function."""
        if node.name not in EXCEL_FUNCTIONS:
            raise ValueError(f"Unknown function: {node.name}")

        return EXCEL_FUNCTIONS[node.name](
            *(self._evaluate_node(arg, current_sheet) for arg in node.arguments)
        )

    def _evaluate_name_ref(
        self, node: NameReference, current_sheet: str | Worksheet
    ) -> ExcelValue:
        """Evaluate a named reference."""
        sheet_name = self._get_sheet(current_sheet).title
        # Sheet-level name first
        name_ref = self.defined_names.get(sheet_name, {}).get(node.name.upper())
        # Workbook-level name second
        if name_ref is None:
            name_ref = self.defined_names.get("", {}).get(node.name.upper())
        if name_ref is None:
            raise NameNotFound(f"Undefined name: {node.name}")
        
        if isinstance(name_ref, CellReference):
            #if the result in in the cache, return it
            cache_key = (name_ref.sheet, name_ref.row, column_as_int(name_ref.column))
            if cache_key in self.cache:
                return self.cache[cache_key]

        if self.evaluation_stack.contains(node, current_sheet):
            cycle_path = self.evaluation_stack.format_cycle_path(node)
            raise CycleError(f"Detected cycle: {cycle_path}")

        self.evaluation_stack.push(node, current_sheet)
        result = self._evaluate_node(name_ref, current_sheet)
        self.evaluation_stack.pop()
        return result

    def _apply_array_formula(
        self, formula: ArrayFormula, current_sheet: str | Worksheet
    ):
        min_col, min_row, max_col, max_row = range_boundaries(formula.ref)
        assert isinstance(formula.text, str), (
            f"Expected array formula string, received: {formula.text}"
        )

        # Ensure all boundaries are not None
        assert (
            min_col is not None
            and min_row is not None
            and max_col is not None
            and max_row is not None
        ), f"Invalid range for array formula: {formula.ref}"

        # Calculate dimensions of the target range
        range_width = max_col - min_col + 1
        range_height = max_row - min_row + 1
        range_shape = (range_height, range_width)

        result = self.evaluate(formula.text, current_sheet)
        
        # Handle array formula results
        if isinstance(result, list):
            # Get the shape of the result and validate it's regular
            result_shape = array_shape(result)

            # Handle special case: transpose 1D horizontal array to vertical if target range is vertical
            if result_shape[0] == 1 and range_shape[0] > 1 and range_shape[1] == 1:
                result = [[val] for val in result]  # Convert to vertical array
                result_shape = (len(result), 1)

            # Validate result dimensions don't exceed target range
            if result_shape[0] > range_shape[0] or result_shape[1] > range_shape[1]:
                raise ValueError(
                    f"Array formula result shape {result_shape} exceeds "
                    f"target range shape {range_shape}"
                )

            # Get sheet info for caching
            sheet = self._get_sheet(current_sheet)
            sheet_name = sheet.title

            # Cache array results based on dimensionality
            if result_shape[0] == 1:
                # Handle 1D horizontal array (single row)
                row_data = result[0] if isinstance(result[0], list) else result
                for j, value in enumerate(row_data):
                    value = "#N/A" if not value else value
                    self.cache[(sheet_name, min_row, min_col + j)] = value
            else:
                # Handle 2D array or 1D vertical array
                for i in range(result_shape[0]):
                    row = min_row + i
                    row_data = result[i]

                    if not isinstance(row_data, list):
                        # Single value in vertical array
                        self.cache[(sheet_name, row, min_col)] = row_data
                    else:
                        # Row of values
                        for j, value in enumerate(row_data):
                            value = "#N/A" if not value else value
                            self.cache[(sheet_name, row, min_col + j)] = value

        # Handle scalar (non-array) results
        elif isinstance(result, (float, int)):
            sheet = self._get_sheet(current_sheet)
            self.cache[(sheet.title, min_row, min_col)] = result

    def load_defined_names(self) -> None:
        """Load defined names from the workbook."""
        # Load workbook-level names
        for name in self.workbook.defined_names.values():
            name_def = self._parse_name_definition(name)
            if name_def is not None:
                self.defined_names[""][name.name.upper()] = name_def

        # Load sheet-level names
        for sheet in self.workbook.worksheets:
            if not sheet.defined_names:
                continue
            self.defined_names[sheet.title] = {}
            for name in sheet.defined_names.values():
                name_def = self._parse_name_definition(name)
                if name_def is not None:
                    self.defined_names[sheet.title][name.name.upper()] = name_def

    def load_array_formulae(self):
        for ws in self.workbook.worksheets:
            for cell, affected_range in ws.array_formulae.items():
                min_col, min_row, max_col, max_row = range_boundaries(affected_range)
                assert (
                    min_col is not None
                    and min_row is not None
                    and max_col is not None
                    and max_row is not None
                ), (
                    f"Invalid range for array formula in sheet {ws.title}: {affected_range}"
                )
                source_row, source_col = coordinate_to_tuple(cell)
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        self.array_formulae[(ws.title, row, col)] = (
                            source_row,
                            source_col,
                        )

    def _parse_name_definition(
        self, name: DefinedName
    ) -> CellRange | CellReference | None:
        """Parse a defined name into an AST node using its definitions."""
        if name.value == "#REF!":
            return None
        for sheet, coord in name.destinations:
            # We only need the first definition
            # Remove absolute references
            ref = coord.replace("$", "")

            if ":" in ref:  # Range reference
                start, end = ref.split(":")
                return CellRange(
                    start=self._parse_cell_ref(start, sheet),
                    end=self._parse_cell_ref(end, sheet),
                )
            else:  # Single cell reference
                return self._parse_cell_ref(ref, sheet)

        raise NameNotFound(f"No valid definition found for name: {name.name}")

    def _parse_cell_ref(self, ref: str, sheet: str | None = None) -> CellReference:
        """Parse a cell reference string into a CellReference node."""
        # Simple implementation - assumes valid references
        col = "".join(c for c in ref if c.isalpha())
        row = int("".join(c for c in ref if c.isdigit()))
        return CellReference(column=col, row=row, sheet=sheet)

    def _get_sheet(self, sheet: Worksheet | str) -> Worksheet:
        if isinstance(sheet, str):
            if sheet not in self.workbook:
                raise SheetNotFound(f'Worksheet "{sheet}" not found.')
            return self.workbook[sheet]
        else:
            return sheet

    def invalidate_cell(self, sheet: str, row: int, col: Union[int, str]) -> None:
        """Invalidate a cell in the cache."""
        self.cache.pop((sheet, column_as_int(col), row), None)

    def clear_cache(self) -> None:
        """Clear the entire cache."""
        self.cache.clear()

    def _evaluate_array(
        self, node: ArrayLiteral, current_sheet: str | Worksheet
    ) -> ExcelValue:
        """Evaluate an array literal, returning an array value."""
        elements = [self._evaluate_node(elem, current_sheet) for elem in node.elements]

        if node.vertical:
            # For vertical arrays, wrap each element in its own array
            return [[elem] for elem in elements]

        return elements
