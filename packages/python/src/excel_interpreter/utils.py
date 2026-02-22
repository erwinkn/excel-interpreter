import math
import re
from pathlib import Path
from typing import IO, Literal, Sequence, cast, overload

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import excel_interpreter.ast as ast
from excel_interpreter.types import ExcelValue

from .expression import (
    Cell,
    Constant,
    Expression,
    ExprOrValue,
)

# Constants
CELL_REF_REGEX = re.compile(r"\$?([A-Z]+)\$?(\d+)$")


def extract_cell_reference(
    ref: str, sheet: str | None = None
) -> ast.CellReference | None:
    """Parse a cell reference, returning None if invalid."""
    match = CELL_REF_REGEX.match(ref)
    if match:
        col, row = match.groups()
        col_absolute = "$" in ref and ref.index("$") == 0
        row_absolute = "$" in ref and (
            len(ref.split("$")) > 2  # Has two $ signs
            or (ref.index("$") > 0)  # Has one $ sign not at start
        )
        return ast.CellReference(
            column=col,
            row=int(row),
            sheet=sheet,
            absolute_col=col_absolute,
            absolute_row=row_absolute,
        )
    return None


def pretty_print_ast(node: ast.ASTNode, indent: int = 0) -> None:
    """Print AST in a human-readable format."""
    indent_str = "  " * indent
    if isinstance(node, ast.ExcelFunction):
        print(f"{indent_str}Function: {node.name}")
        for i, arg in enumerate(node.arguments):
            print(f"{indent_str}  Argument {i + 1}:")
            pretty_print_ast(arg, indent + 2)
    elif isinstance(node, ast.BinaryOperation):
        print(f"{indent_str}Binary Operation: {node.operator}")
        print(f"{indent_str}  Left:")
        pretty_print_ast(node.left, indent + 2)
        print(f"{indent_str}  Right:")
        pretty_print_ast(node.right, indent + 2)
    elif isinstance(node, ast.UnaryOperation):
        print(f"{indent_str}Unary Operation: {node.operator}")
        pretty_print_ast(node.operand, indent + 1)
    elif isinstance(node, ast.CellReference):
        sheet_str = f"'{node.sheet}'!" if node.sheet else ""
        print(f"{indent_str}Cell Reference: {sheet_str}{node.column}{node.row}")
    elif isinstance(node, ast.CellRange):
        print(f"{indent_str}Cell Range:")
        print(f"{indent_str}  Start:")
        pretty_print_ast(node.start, indent + 2)
        print(f"{indent_str}  End:")
        pretty_print_ast(node.end, indent + 2)
    elif isinstance(node, ast.Constant):
        print(f"{indent_str}Constant: {node.value}")
    elif isinstance(node, ast.NameReference):
        print(f"{indent_str}Name Reference: {node.name}")
    else:
        print(f"{indent_str}Unknown node type: {type(node)}")


def _pretty_print_recursive(node: ast.ASTNode, indent: int = 0):
    indent_str = "  " * indent

    if isinstance(node, ast.ExcelFunction):
        print(f"{node.name}", end="")  # Print name (caller handles initial indent)
        if node.arguments:
            print()  # Newline after name
            print(indent_str + "(", end="")  # Open paren on new line, same indent
            print()  # Newline after open paren
            for i, arg in enumerate(node.arguments):
                print("  " * (indent + 1), end="")  # Indent for argument line
                _pretty_print_recursive(arg, indent + 1)  # Print argument content
                if i < len(node.arguments) - 1:
                    print(",", end="")  # Comma after argument (no newline yet)
                print()  # Newline after argument (and comma)
            print(
                indent_str + ")", end=""
            )  # Closing paren on new line, original indent
        else:
            print("()", end="")  # No arguments: NAME()

    elif isinstance(node, ast.BinaryOperation):
        # Let operands handle their own formatting based on type
        _pretty_print_recursive(node.left, indent)
        print(f" {node.operator} ", end="")
        _pretty_print_recursive(node.right, indent)

    elif isinstance(node, ast.UnaryOperation):
        print(f"{node.operator}", end="")
        _pretty_print_recursive(node.operand, indent)

    elif isinstance(node, ast.CellReference):
        sheet_str = f"'{node.sheet}'!" if node.sheet else ""
        col_prefix = "$" if node.absolute_col else ""
        row_prefix = "$" if node.absolute_row else ""
        print(f"{sheet_str}{col_prefix}{node.column}{row_prefix}{node.row}", end="")

    elif isinstance(node, ast.CellRange):
        _pretty_print_recursive(node.start, indent)
        print(":", end="")
        _pretty_print_recursive(node.end, indent)

    elif isinstance(node, ast.Constant):
        if isinstance(node.value, str):
            print(f'"{node.value}"', end="")
        elif isinstance(node.value, bool):
            print(str(node.value).upper(), end="")
        else:
            print(f"{node.value}", end="")

    elif isinstance(node, ast.NameReference):
        print(f"{node.name}", end="")

    else:
        # Fallback for unknown nodes - print without indent
        print(f"<?{type(node).__name__}?>", end="")


def pretty_print_formula(formula_or_ast: str | ast.ASTNode) -> None:
    """Print a formula string or AST in a human-readable format with indentation."""
    node: ast.ASTNode

    if isinstance(formula_or_ast, str):
        try:
            # Avoid circular import, keep imports local
            from excel_interpreter.parser import ExcelParser
            from excel_interpreter.tokenizer import ExcelTokenizer

            # Remove leading = if present for parsing
            formula_str = formula_or_ast.lstrip("=")
            node = ExcelParser(ExcelTokenizer(formula_str).tokenize()).parse()
        except Exception as e:
            print(f"Error parsing formula: {e}")
            print(formula_or_ast)  # Print the original string if parsing fails
            return
    else:
        node = formula_or_ast

    _pretty_print_recursive(node, 0)
    print()  # Ensure final newline


"Simple type helper to interoperate between regular numbers and Excel expressions"
NumberOrExpr = Expression | float


def copy_expr(expr: NumberOrExpr) -> NumberOrExpr:
    if isinstance(expr, Expression):
        return expr.copy()
    else:
        return expr


# The overload are here to provide better type hints when using this function.
# They do not affect behavior so do not worry about them, you can jump to the
# main function definition.
@overload
def read_str_cell(
    ws: Worksheet,
    row: int,
    col: int,
    *,
    default: str | None = None,
    optional: Literal[False] = False,
) -> str: ...


@overload
def read_str_cell(
    ws: Worksheet,
    row: int,
    col: int,
    *,
    # Optional means no default
    default: None = None,
    optional: Literal[True],
) -> str | None: ...


def read_str_cell(
    ws: Worksheet,
    row: int,
    col: int,
    *,
    default: str | None = None,
    optional: bool = False,
) -> str | None:
    "Helper to read a cell and ensure the return type is as expected"

    value = ws.cell(row, col).value
    if not value:
        if default is None and not optional:
            raise Exception(
                f"Cell {get_column_letter(col)}{row} is empty (worksheet {ws.title})"
            )
        return default
    else:
        return str(value)


# The overload are here to provide better type hints when using this function.
# They do not affect behavior so do not worry about them, you can jump to the
# main function definition.
@overload
def read_float_cell(
    ws: Worksheet,
    row: int,
    col: int,
    *,
    default: float | None = None,
    optional: Literal[False] = False,
) -> float: ...


@overload
def read_float_cell(
    ws: Worksheet,
    row: int,
    col: int,
    *,
    # Optional means no default
    default: None = None,
    optional: Literal[True],
) -> float | None: ...


def read_float_cell(
    ws: Worksheet,
    row: int,
    col: int,
    *,
    default: float | None = None,
    optional: bool = False,
) -> float | None:
    "Helper to read a cell and ensure the return type is as expected"
    value = ws.cell(row, col).value

    if isinstance(value, (int, float)):
        return value
    else:
        value = str(value)
        # For cases where you have a value like `= 123`
        if value.startswith("="):
            value = value[1:].strip()
        # Try parsing the value into a float
        try:
            float_value = float(str(value))
            if not math.isnan(float_value):
                return float_value
        except:
            ...
        # Fall through for all failed conversions
        if default is None and not optional:
            if len(value.strip()) == 0:
                raise Exception(
                    f"Cell {get_column_letter(col)}{row} is empty (worksheet {
                        ws.title
                    })"
                )
            else:
                raise Exception(
                    f"Cell {get_column_letter(col)}{row} contains a non-numeric value: {
                        value
                    } (worksheet {ws.title})"
                )
        return default


# The overload are here to provide better type hints when using this function.
# They do not affect behavior so do not worry about them, you can jump to the
# main function definition.
@overload
def read_int_cell(
    ws: Worksheet,
    row: int,
    col: int,
    *,
    default: int | None = None,
    optional: Literal[False] = False,
) -> int: ...


@overload
def read_int_cell(
    ws: Worksheet,
    row: int,
    col: int,
    *,
    # Optional means no default
    default: None = None,
    optional: Literal[True],
) -> int | None: ...


def read_int_cell(
    ws: Worksheet,
    row: int,
    col: int,
    *,
    default: int | None = None,
    optional: bool = False,
) -> int | None:
    "Helper to read a cell and ensure the return type is as expected"

    value = str(ws.cell(row, col).value)

    if isinstance(value, (int, float)):
        return int(value)
    else:
        value = str(value)
        # For cases where you have a value like `= 123`
        if value.startswith("="):
            value = value[1:].strip()
        # Try parsing the value into an integer
        try:
            int_value = int(value)
            if not math.isnan(int_value):
                return int_value
        except:
            ...
        if default is None and not optional:
            if len(value.strip()) == 0:
                raise Exception(
                    f"Cell {get_column_letter(col)}{row} is empty (worksheet {
                        ws.title
                    })"
                )
            else:
                raise Exception(
                    f"Cell {get_column_letter(col)}{row} contains a non-integer value: {
                        value
                    } (worksheet {ws.title})"
                )
        return default


def add_style(
    expr: ExprOrValue,
    *,
    number_format: str | None = None,
    font: Font | None = None,
    fill: PatternFill | None = None,
    border: Border | None = None,
    alignment: Alignment | None = None,
    protection: Protection | None = None,
):
    if not isinstance(expr, Expression):
        expr = Constant(expr)

    return expr.style(
        number_format=number_format,
        font=font,
        fill=fill,
        border=border,
        alignment=alignment,
        protection=protection,
    )


def number_format(expr: ExprOrValue, number_format: str):
    if not isinstance(expr, Expression):
        expr = Constant(expr)
    return expr.number_format(number_format)


def fill(expr: ExprOrValue, fill: PatternFill):
    if not isinstance(expr, Expression):
        expr = Constant(expr)
    return expr.style(fill=fill)


def ensure_is_expr(value: int | float | Expression):
    if not isinstance(value, Expression):
        return Constant(value)
    else:
        return value


def get_cell_column(ws: Worksheet, column: str | int, start_row: int, end_row: int):
    assert end_row >= start_row, (
        f"End row comes before start row: {end_row} (end) < {start_row} (start)"
    )
    cells: list[Cell] = []
    for row in range(start_row, end_row + 1):
        cells.append(Cell(ws, row, column))
    return cells


def get_cell_row(ws: Worksheet, row: int, start_col: int | str, end_col: int | str):
    start_idx = (
        column_index_from_string(start_col) if isinstance(start_col, str) else start_col
    )
    end_idx = column_index_from_string(end_col) if isinstance(end_col, str) else end_col

    assert end_idx >= start_idx, (
        f"End column comes before start column: {end_col} (end) < {start_col} (start)"
    )
    cells: list[Cell] = []
    for col in range(start_idx, end_idx + 1):
        cells.append(Cell(ws, row, col))
    return cells


def get_or_replace_sheet(
    wb: Workbook, title: str, delete: bool, idx: int | None = None
) -> Worksheet:
    if title in wb:
        if delete:
            ws = wb[title]
            idx = idx or wb.index(ws)
            wb.remove(ws)
            return wb.create_sheet(title, idx)
        else:
            return wb[title]
    else:
        return wb.create_sheet(title, idx)


def column_as_int(col: int | str):
    if isinstance(col, str):
        col = column_index_from_string(col)
    return col


def column_as_str(col: int | str):
    if isinstance(col, int):
        col = get_column_letter(col)
    return col


def extract_sub_table_as_df(
    file: str | Path | IO[bytes],
    sheet_name: str,
    usecols: str | Sequence[str] | int | Sequence[int],
    start_row: int,
    end_row: int,
):
    table = pd.read_excel(
        file,
        sheet_name,
        usecols=usecols,
        skiprows=start_row - 1,
        nrows=end_row - start_row + 1,
    )
    return table


def write_to_excel_and_size_columns(
    df: pd.DataFrame,
    writer: pd.ExcelWriter,
    sheet_name: str,
    na_rep: str = "N/A",
    column_sizes: dict[str, float] | None = None,
    index=False,
    freeze_panes: tuple[int, int] | None = None,
    columns: list[str] | None = None,
) -> None:
    """Requires using `engine='openpyxl'` for `pandas.ExcelWriter`"""
    if column_sizes is None:
        column_sizes = {}
    df.to_excel(
        writer,
        sheet_name=sheet_name,
        na_rep=na_rep,
        index=index,
        columns=columns,
        freeze_panes=freeze_panes,
    )

    # If DataFrame has a multi-index
    if isinstance(df.index, pd.MultiIndex):
        index_columns = df.index.levels
    else:
        index_columns = [df.index.to_series()]

    # 1-based indexing
    i = 1

    if index:
        for index_column in index_columns:
            index_name = cast(str, index_column.name if index_column.name else "index")
            col_size = column_sizes.get(
                index_name, index_column.astype(str).str.len().max()
            )
            writer.sheets[sheet_name].column_dimensions[
                get_column_letter(i)
            ].width = float(col_size)
            i += 1

    for col in df.columns:
        col_size = column_sizes.get(col)
        if col_size is None:
            # max content length
            col_size = df[col].astype(str).str.len().max()
            # also consider header
            col_size = max(col_size, len(col))

        col_letter = get_column_letter(i)
        writer.sheets[sheet_name].column_dimensions[col_letter].width = float(col_size)

        i += 1


def auto_resize_columns(
    ws: Worksheet, min_width: float = 0, max_width: float = 50
) -> None:
    """Automatically resize columns in a worksheet based on content.

    Args:
        ws: The worksheet to resize columns in
        min_width: Minimum column width (default 0)
        max_width: Maximum column width (default 50)
    """
    for col in ws.columns:
        # Get max length of column content
        max_length = 0

        # Check header and all cell values
        for cell in col:
            try:
                if cell.value:
                    # Get length of cell value as string
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        # Set column width with constraints
        adjusted_width = max(min_width, min(max_length + 2, max_width))
        column = col[0].column_letter  # Get column letter
        ws.column_dimensions[column].width = adjusted_width


def array_shape(array: ExcelValue) -> tuple[int, int]:
    """
    Determine the shape of an array and validate it has a regular structure.

    Returns a tuple of (height, width):
    - For 1D arrays: always returns (1, n)
    - For 2D arrays: returns (rows, columns)

    Raises ValueError if the array has irregular/jagged rows.
    """
    if not isinstance(array, list):
        return (1, 1)
    if not array:
        return (0, 0)

    # Check if it's a 2D array (at least one element is a list)
    is_2d = any(isinstance(item, list) for item in array)

    if not is_2d:
        # It's a 1D array - in Excel these are always horizontal (row vectors)
        return (1, len(array))

    # It's a 2D array, validate all rows have the same length
    heights = len(array)

    # Get the width of each row
    row_widths = []
    for row in array:
        if isinstance(row, list):
            row_widths.append(len(row))
        else:
            # Handle scalar values in a 2D array
            row_widths.append(1)

    # Check if all rows have the same width
    if len(set(row_widths)) > 1:
        raise ValueError(f"Array has irregular row lengths: {row_widths}")

    # Return the shape as (height, width)
    return (heights, row_widths[0] if row_widths else 0)
