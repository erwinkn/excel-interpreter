from abc import ABC, abstractmethod
from contextlib import contextmanager
import logging
import re
from typing import (
    Any,
    Generic,
    Literal,
    NamedTuple,
    TypeVar,
    Union,
    cast,
    overload,
)
from typing_extensions import Self
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection
from openpyxl.cell import Cell as OpenpyxlCell
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.comments import Comment


ExcelAllowedValue = Union[int, float, str, bool]
ExprOrValue = Union[ExcelAllowedValue, "Expression"]

default_author: str | None = None

registration_key = "__expressions__"

fast_mode = False


def enable_fast_excel_formulas():
    global fast_mode
    fast_mode = True


def disable_fast_excel_formulas():
    global fast_mode
    fast_mode = False


def set_default_author(author: str | None):
    global default_author
    default_author = author


def format_within_formula(expr: ExprOrValue, ws: Worksheet):
    # Special cases for raw strings used in a formula: they need to be wrapped in quotes
    if isinstance(expr, str):
        return '"' + expr + '"'
    elif (
        isinstance(expr, Constant)
        and isinstance(expr.value, str)
        and not expr.has_been_written
    ):
        return '"' + expr.value + '"'
    # Normal case
    if isinstance(expr, Expression):
        return expr.formula_or_ref(ws)
    else:
        return str(expr)


def parenthesize(expr: str):
    if expr.startswith("(") and expr.endswith(")"):
        # Need to check we're not in a case like "(A+2)*(B+3)"
        parenthesis_counter = 1
        for ch in expr[1:-1]:
            if ch == "(":
                parenthesis_counter += 1
            elif ch == ")":
                parenthesis_counter -= 1
            # That means the initial parenthesis was closed before the end of
            # the expression -> it's not already parenthesized.
            if parenthesis_counter == 0:
                return "(" + expr + ")"
        return expr
    return "(" + expr + ")"


class ExpressionLocation(NamedTuple):
    worksheet: Worksheet
    row: int
    col: int


class Expression(ABC):
    # All the styling options available for individual cells in `openpyxl`
    # https://openpyxl.readthedocs.io/en/stable/styles.html
    number_format_value: str | None = None
    font: Font | None = None
    fill: PatternFill | None = None
    border: Border | None = None
    alignment: Alignment | None = None
    protection: Protection | None = None
    comment_value: Comment | None = None

    locations: list[ExpressionLocation]
    main_location_idx: int = 0

    def __init__(self):
        self.id = id(self)
        self.locations = []
        self.has_been_written = False

    @abstractmethod
    def deps(self) -> list[ExprOrValue] | tuple[ExprOrValue, ...]: ...

    @abstractmethod
    def copy(self) -> Self: ...

    @abstractmethod
    def __formula__(self, current_worksheet: Worksheet) -> str:
        "Returns the full expression represented by the object, without simplification."
        ...

    def __simplify__(self) -> ExprOrValue:
        return self

    def simplify(self, nested=False):
        if nested and self.has_been_written:
            return self

        global simplification_cache

        if simplification_cache and self.id in simplification_cache:
            return simplification_cache[self.id]

        # TODO: create simplification cache if doesn't already exist?
        # -> only if we notice slowdown due to repeated simplifications of
        # highly nested formulas

        simpl = self.__simplify__()
        if simplification_cache:
            simplification_cache[self.id] = simpl
        return simpl

    def formula(self, current_worksheet: Worksheet):
        expr = self.simplify()
        if isinstance(expr, Expression):
            return expr.__formula__(current_worksheet)
        else:
            return str(expr)

    @property
    def main_location(self):
        assert self.has_been_written, "Expression has no location yet"
        return self.locations[self.main_location_idx]

    def ref(self, current_worksheet: Worksheet):
        assert self.has_been_written, (
            "Called Expression.ref() on an expression that has not yet been written, impossible to provide a reference. Please use `formula_or_ref()`."
        )
        loc = self.locations[self.main_location_idx]
        # Try to find a location on the same worksheet
        if loc.worksheet != current_worksheet:
            for candidate in self.locations:
                if candidate.worksheet == current_worksheet:
                    loc = candidate
                    break
        ref = f"{get_column_letter(loc.col)}{loc.row}"
        if loc.worksheet != current_worksheet:
            ref = f"'{loc.worksheet.title}'!{ref}"
        return ref

    def formula_or_ref(self, current_worksheet: Worksheet) -> str:
        "Returns the coordinate of the last write of this expression, or its complete formula if it has not been written yet."
        if self.has_been_written:
            return self.ref(current_worksheet)
        else:
            formula = self.__formula__(current_worksheet)
            # Special case for constant strings used in a fomula
            # if isinstance(self, Constant) and isinstance(self.value, str):
            #     formula = '"' + formula + '"'
            return formula

    def update_location(
        self,
        i: int,
        *,
        # All the styling options available for individual cells in `openpyxl`
        # https://openpyxl.readthedocs.io/en/stable/styles.html
        number_format: str | None = None,
        font: Font | None = None,
        fill: PatternFill | None = None,
        border: Border | None = None,
        alignment: Alignment | None = None,
        protection: Protection | None = None,
        comment: Comment | None = None,
        set_as_main_location=False,
    ):
        # Cells shouldn't update their source location
        if isinstance(self, Cell) and i == 0:
            return
        loc = self.locations[i]
        cell = loc.worksheet.cell(loc.row, loc.col)
        if not fast_mode:
            if i == self.main_location_idx:
                # Constants that have not yet been written (= we're writing the raw value, not a reference)
                # are special case. First, we don't add the '=' sign, since they don't represent a formula.
                # Also, for numeric values, we want to store them as numeric values, not text,
                # for formatting + calculation purposes.
                if isinstance(self, Constant):
                    cell.value = self.value
                else:
                    cell.value = "= " + self.formula(loc.worksheet)
            else:
                # Do not use .ref(), it can easily create a self reference, as it looks for a ref in the same worksheet.
                main_loc = self.locations[self.main_location_idx]
                ref = f"{get_column_letter(main_loc.col)}{main_loc.row}"
                if loc.worksheet != main_loc.worksheet:
                    ref = f"'{main_loc.worksheet.title}'!{ref}"
                cell.value = "= " + ref

        if number_format:
            cell.number_format = number_format
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if border:
            cell.border = border
        if alignment:
            cell.alignment = alignment
        if protection:
            cell.protection = protection
        if comment:
            cell.comment = comment

        if set_as_main_location:
            self.main_location_idx = len(self.locations)

    def write(
        self,
        ws: Worksheet,
        row: int,
        col: int | str,
        *,
        # All the styling options available for individual cells in `openpyxl`
        # https://openpyxl.readthedocs.io/en/stable/styles.html
        number_format: str | None = None,
        font: Font | None = None,
        fill: PatternFill | None = None,
        border: Border | None = None,
        alignment: Alignment | None = None,
        protection: Protection | None = None,
        comment: Comment | None = None,
        set_as_main_location=False,
    ):
        "Write to an Excel cell and save the location for future references. Optionally, override the default Expression styling."
        if isinstance(col, str):
            col = column_index_from_string(col)

        cell = ws.cell(row, col)

        self.locations.append(ExpressionLocation(ws, row, col))
        self.update_location(
            len(self.locations) - 1,
            number_format=number_format or self.number_format_value,
            font=font or self.font,
            border=border or self.border,
            fill=fill or self.fill,
            alignment=alignment or self.alignment,
            protection=protection or self.protection,
            comment=comment or self.comment_value,
            set_as_main_location=set_as_main_location,
        )

        # Register this expression with the workbook. This can be used to
        # simplify expressions through cross-references later on.
        register_expression(ws.parent, self)

        # Important: only do this after the actual write to the location.
        # Otherwise, the call to `self.simplify()` when building the formula
        # will return early.
        self.has_been_written = True

        return cell

    def style(
        self,
        *,
        number_format: str | None = None,
        font: Font | None = None,
        fill: PatternFill | None = None,
        border: Border | None = None,
        alignment: Alignment | None = None,
        protection: Protection | None = None,
    ):
        "Set styling options for this cell. Will only affect subsequent writes, not previous ones."
        if number_format is not None:
            self.number_format_value = number_format
        if font is not None:
            self.font = font
        if fill is not None:
            self.fill = fill
        if border is not None:
            self.border = border
        if alignment is not None:
            self.alignment = alignment
        if protection is not None:
            self.protection = protection

        return self

    def number_format(self, value: str):
        self.number_format_value = value
        return self

    # Those values for width and height are the default ones
    def comment(
        self,
        text: str,
        author: str | None = None,
        *,
        height: int = 79,
        width: int = 144,
    ):
        author = author or default_author
        if author is None:
            if default_author is None:
                raise Exception(
                    "No author specified for Excel comment. If you don't want to specify the author every time, use the `excel_interpreter.set_default_author` function to set it once."
                )
            author = default_author

        self.comment_value = Comment(
            text=text, author=author, height=height, width=width
        )

    def evaluate(self):
        return evaluate_expression(self)

    def __pos__(self):
        # no effect, just skip for more concise formula?
        return self

    def __neg__(self):
        return UnaryExpression("-", self)

    def __add__(self, other: "int | float | Expression"):
        return BinaryExpression(self, "+", other)

    def __radd__(self, other: "int | float | Expression"):
        return BinaryExpression(other, "+", self)

    def __sub__(self, other: "int | float | Expression"):
        return BinaryExpression(self, "-", other)

    def __rsub__(self, other: "int | float | Expression"):
        return BinaryExpression(other, "-", self)

    def __mul__(self, other: "int | float | Expression"):
        return BinaryExpression(self, "*", other)

    def __rmul__(self, other: "int | float | Expression"):
        return BinaryExpression(other, "*", self)

    def __truediv__(self, other: "int | float | Expression"):
        return BinaryExpression(self, "/", other)

    def __rtruediv__(self, other: "int | float | Expression"):
        return BinaryExpression(other, "/", self)

    def __pow__(self, other: "int | float | Expression"):
        return ExcelFunction("POWER", self, other)

    def __rpow__(self, other: "int | float | Expression"):
        return ExcelFunction("POWER", other, self)

    def __gt__(self, other: "int | float | Expression"):
        return BinaryExpression(self, ">", other)

    def __lt__(self, other: "int | float | Expression"):
        return BinaryExpression(self, "<", other)

    def __ge__(self, other: "int | float | Expression"):
        return BinaryExpression(self, ">=", other)

    def __le__(self, other: "int | float | Expression"):
        return BinaryExpression(self, "<=", other)

    def __eq__(self, other: "int | float | Expression"):
        return BinaryExpression(self, "=", other)

    def __ne__(self, other: "int | float | Expression"):
        return BinaryExpression(self, "<>", other)


ConstantValue = TypeVar("ConstantValue", bound=ExcelAllowedValue)


class Constant(Expression, Generic[ConstantValue]):
    value: ConstantValue

    def __init__(self, value: ConstantValue):
        super().__init__()
        self.value = value

    def __formula__(self, current_worksheet: Worksheet) -> str:
        return str(self.value)

    def deps(self) -> list | tuple:
        return ()

    def copy(self) -> "Constant[ConstantValue]":
        return Constant(self.value)

    def __str__(self):
        return f"Constant({self.value})"


UnaryOp = Literal["+", "-"]


class UnaryExpression(Expression):
    op: UnaryOp
    expr: ExprOrValue

    def __init__(self, op: UnaryOp, expr: ExprOrValue):
        super().__init__()
        self.op = op
        self.expr = expr

    def __formula__(self, current_worksheet: Worksheet) -> str:
        expr = format_within_formula(self.expr, current_worksheet)
        if self.op == "+":
            return expr

        if isinstance(self.expr, BinaryExpression):
            expr = "(" + expr + ")"
        return self.op + expr

    def __simplify__(self) -> ExprOrValue:
        if isinstance(self.expr, Expression):
            simpl = self.expr.simplify()
            if simpl is not self.expr:
                return UnaryExpression(self.op, simpl)
        return self

    def deps(self) -> list | tuple:
        return (self.expr,)

    def copy(self) -> "UnaryExpression":
        return UnaryExpression(self.op, self.expr)

    def __str__(self) -> str:
        return f"UnaryExpression({self.op}, {self.expr})"


BinaryOp = Literal["+", "-", "*", "/", "<", ">", "<=", ">=", "=", "<>"]


def precedence(op: BinaryOp):
    if op == "+" or op == "-":
        return 1
    elif op == "*" or op == "/":
        return 2
    elif op == "<" or op == "<=" or op == "<>" or op == "=" or op == ">" or op == ">=":
        return 3
    else:
        raise ValueError(f"Unknown binary operation: {op}")


def is_constant(expr: ExprOrValue, constant: int | float):
    if not isinstance(expr, Expression):
        return expr == constant
    else:
        return (
            isinstance(expr, Constant)
            and not expr.has_been_written
            and expr.value == constant
        )


def should_parenthesize_left(
    op: "BinaryExpression", left: Expression | int | str | float
):
    if (
        isinstance(left, BinaryExpression)
        and not left.has_been_written
        and precedence(op.op) > precedence(left.op)
    ):
        return True
    else:
        return False


def should_parenthesize_right(
    op: "BinaryExpression", right: Expression | int | str | float
):
    # Parenthesis for the right member is a bit more complicated,
    # we can't rely on precedence alone, due to two cases:
    # - subtraction
    # - division
    if isinstance(right, BinaryExpression) and not right.has_been_written:
        # Those 3 cases are mutually exclusive, so there's no risk of adding parenthesis twice
        if precedence(op.op) > precedence(right.op):
            return True

        if op.op == "-" and (right.op == "+" or right.op == "-"):
            return True

        if op.op == "/" and (right.op == "*" or right.op == "/"):
            return True

    return False


class BinaryExpression(Expression):
    left: ExprOrValue
    op: BinaryOp
    right: ExprOrValue

    def __init__(self, left: ExprOrValue, op: BinaryOp, right: ExprOrValue) -> None:
        super().__init__()
        self.op = op
        self.left = left
        self.right = right

    def deps(self) -> list | tuple:
        return (self.left, self.right)

    def __simplify__(self):
        left = self.left
        if isinstance(left, Expression):
            left = left.simplify(nested=True)

        right = self.right
        if isinstance(right, Expression):
            right = right.simplify(nested=True)

        if self.op == "+":
            if is_constant(left, 0):
                return right
            if is_constant(right, 0):
                return left
        if self.op == "-":
            if is_constant(left, 0):
                return UnaryExpression("-", right)
            if is_constant(right, 0):
                return left
        if self.op == "*":
            if is_constant(left, 0):
                return left
            if is_constant(right, 1):
                return left
            if is_constant(left, 1):
                return right
            if is_constant(right, 0):
                return right
        if self.op == "/":
            if is_constant(left, 0):
                return left
            if is_constant(right, 1):
                return left

        if left is not self.left or right is not self.right:
            return BinaryExpression(left, self.op, right)
        else:
            return self

    def __formula__(self, current_worksheet: Worksheet) -> str:
        # Assume formula has already been simplified
        left = format_within_formula(self.left, current_worksheet)
        right = format_within_formula(self.right, current_worksheet)

        if should_parenthesize_left(self, self.left):  # reference the expr here
            left = "(" + left + ")"
        if should_parenthesize_right(self, self.right):  # reference the expr here
            right = "(" + right + ")"

        return f"{left} {self.op} {right}"

    def copy(self) -> "BinaryExpression":
        return BinaryExpression(self.left, self.op, self.right)

    def __str__(self) -> str:
        return f'BinaryExpression({self.left}, "{self.op}", {self.right})'


class ExcelFunction(Expression):
    symbol: str
    arguments: tuple[ExprOrValue, ...]

    def __init__(
        self,
        symbol: str,
        *args: ExprOrValue,
    ):
        super().__init__()
        self.symbol = symbol
        self.arguments = args

    def deps(self) -> list | tuple:
        return self.arguments

    def __formula__(self, current_worksheet: Worksheet) -> str:
        arguments = ", ".join(
            format_within_formula(arg, current_worksheet) for arg in self.arguments
        )
        return f"{self.symbol}({arguments})"

    def __simplify__(self) -> ExprOrValue:
        new_args = [
            arg.simplify() if isinstance(arg, Expression) else arg
            for arg in self.arguments
        ]
        if any(
            new_args[i] is not self.arguments[i] for i in range(len(self.arguments))
        ):
            return ExcelFunction(self.symbol, *new_args)
        return self

    def copy(self) -> "ExcelFunction":
        # We can't use `dataclasses.replace` here, as this class it not a dataclass
        # (as we wanted a custom `__init__` method)
        return ExcelFunction(
            self.symbol,
            *self.arguments,
        )

    def __str__(self) -> str:
        args = ", ".join([str(arg) for arg in self.arguments])
        return f"ExcelFunction({self.symbol}, {args})"


class Excel:
    "Umbrella class for Excel functions and helpers"

    @staticmethod
    def sqrt(value: Expression | int | float):
        return ExcelFunction("SQRT", value)

    @staticmethod
    def abs(value: Expression | int | float):
        return ExcelFunction("ABS", value)

    @staticmethod
    def min(*args: ExprOrValue):
        return ExcelFunction("MIN", *args)

    @staticmethod
    def max(*args: ExprOrValue):
        return ExcelFunction("MAX", *args)

    @staticmethod
    def ceil(value: Expression | int | float, significance: Expression | int | float):
        return ExcelFunction("CEILING", value, significance)

    @staticmethod
    def trunc(value: Expression | int | float, num_digits: Expression | int):
        return ExcelFunction("TRUNC", value, num_digits)

    @staticmethod
    def pi():
        return ExcelFunction("PI")

    @staticmethod
    def sum(*args: ExprOrValue):
        return ExcelFunction("SUM", *args)

    @staticmethod
    def IF(
        condition: Expression | bool,
        ok: ExprOrValue,
        not_ok: ExprOrValue,
    ):
        return ExcelFunction("IF", condition, ok, not_ok)

    @staticmethod
    def IFERROR(value: ExprOrValue, value_if_error: ExprOrValue):
        return ExcelFunction("IFERROR", value, value_if_error)

    @staticmethod
    def IFS(
        *conditions: tuple[Expression | bool, ExprOrValue],
        fallback: ExprOrValue | None = None,
    ):
        args = []
        for condition, if_ok in conditions:
            args.append(condition)
            args.append(if_ok)
        if fallback is not None:
            args.append(True)
            args.append(fallback)
        return ExcelFunction("IFS", *args)

    @staticmethod
    def NA():
        return ExcelFunction("NA")

    @staticmethod
    def OR(*args: Expression | bool):
        return ExcelFunction("OR", *args)

    @staticmethod
    def AND(*args: Expression | bool):
        return ExcelFunction("AND", *args)

    @staticmethod
    def NOT(value: Expression | bool):
        return ExcelFunction("NOT", value)

    @staticmethod
    def exp(value: Expression | int | float):
        return ExcelFunction("EXP", value)

    @staticmethod
    def ln(value: Expression | int | float):
        return ExcelFunction("LN", value)

    @staticmethod
    def log10(value: Expression | int | float):
        return ExcelFunction("LOG10", value)

    @staticmethod
    def IRR(*args: Expression | int | float):
        return ExcelFunction("IRR", *args)

    @staticmethod
    def NPV(rate: Expression | float, *args: Expression | int | float):
        return ExcelFunction("NPV", rate, *args)

    @staticmethod
    def concat(*args: Expression | int | float | str):
        return ExcelFunction("CONCATENATE", *args)


class Cell(Expression):
    def __init__(self, worksheet: Worksheet, row: int, col: int | str) -> None:
        super().__init__()
        if isinstance(col, str):
            col = column_index_from_string(col)

        self.has_been_written = True  # by definition
        self.locations.append(ExpressionLocation(worksheet, row, col))

    @classmethod
    def from_ref(cls, worksheet: Worksheet, ref: str):
        match = re.match(r"([A-Z]+)(\d+)", ref)
        if not match:
            raise ValueError(f'Invalid cell reference: "{ref}"')
        column, row = match.groups()
        return Cell(worksheet, int(row), column)

    @property
    def source_loc(self):
        return self.locations[0]

    @property
    def value(self):
        loc = self.locations[self.main_location_idx]
        return loc.worksheet.cell(loc.row, loc.col).value

    @staticmethod
    def from_openpyxl(cell: OpenpyxlCell):
        assert isinstance(cell.parent, Worksheet), (
            f"Error: expected parent of `openpyxl.cell.Cell` to be a `openpyxl.worksheet.worksheet.Worksheet` object, but received a {type(cell.parent)}"
        )

        return Cell(
            row=cell.row,
            col=cell.column,
            worksheet=cell.parent,
        )

    def __formula__(self, current_worksheet: Worksheet) -> str:
        return self.ref(current_worksheet)

    def deps(self) -> list | tuple:
        return ()

    def copy(self) -> "Cell":
        loc = self.locations[self.main_location_idx]
        return Cell(
            row=loc.row,
            col=loc.col,
            worksheet=loc.worksheet,
        )

    def __str__(self) -> str:
        loc = self.locations[self.main_location_idx]
        return f"Cell({loc.worksheet.title}, {loc.row}, {loc.col})"


str_ref_to_coords_regex = re.compile(r"^([a-zA-Z]+)(\d+)$")


def str_ref_to_coords(ref: str):
    "Convert an Excel reference of the format 'AAAXXX', where X is a number, into row and colum numbers."
    match = str_ref_to_coords_regex.match(ref)
    assert match is not None, (
        f'"{ref}" is not a valid Excel reference (expected format: AAAXXX)'
    )
    col, row = match.groups()
    return int(row), column_index_from_string(col)


class Range(Expression):
    start: Expression | str
    end: Expression | str
    range_worksheet: Worksheet | None

    # Can only explicitly specify a worksheet if no Expression is given.
    # Otherwise, the Expression's worksheet is used for reference.
    @overload
    def __init__(self, start: str, end: str, worksheet: Worksheet) -> None: ...

    @overload
    def __init__(self, start: Expression, end: Expression) -> None: ...

    def __init__(
        self,
        start: str | Expression,
        end: str | Expression,
        worksheet: Worksheet | None = None,
    ) -> None:
        super().__init__()
        self.start = start
        self.end = end
        self.worksheet = worksheet
        if isinstance(start, str) and isinstance(end, str) and worksheet is None:
            raise ValueError("Invalid Range construction: missing worksheet")

    def deps(self) -> list | tuple:
        return (self.start, self.end)

    def start_and_end_locations(self):
        start = self.start
        end = self.end
        if isinstance(start, Expression) and isinstance(end, Expression):
            start_loc = start.main_location
            end_loc = end.main_location
        elif isinstance(start, Expression):
            start_loc = start.main_location
            end_row, end_col = str_ref_to_coords(cast(str, end))
            end_loc = ExpressionLocation(start_loc.worksheet, end_row, end_col)
        elif isinstance(end, Expression):
            start_row, start_col = str_ref_to_coords(start)
            end_loc = end.main_location
            start_loc = ExpressionLocation(end_loc.worksheet, start_row, start_col)
        else:
            assert self.worksheet is not None, "Unreachable"
            start_row, start_col = str_ref_to_coords(start)
            start_loc = ExpressionLocation(self.worksheet, start_row, start_col)
            end_row, end_col = str_ref_to_coords(end)
            end_loc = ExpressionLocation(self.worksheet, end_row, end_col)
        return start_loc, end_loc

    def __formula__(self, current_worksheet: Worksheet) -> str:
        if isinstance(self.start, Expression) and not self.start.has_been_written:
            logging.debug("Range error: start expression has not been written")
            return "#REF!"
        if isinstance(self.end, Expression) and not self.end.has_been_written:
            logging.debug("Range error: end expression has not been written")
            return "#REF!"
        start, end = self.start_and_end_locations()
        if start.worksheet != end.worksheet:
            logging.debug(
                "Range error: start and end expressions are in different worksheets."
            )
            return "#REF!"

        ws = start.worksheet
        start = f"{get_column_letter(start.col)}{start.row}"
        end = f"{get_column_letter(end.col)}{end.row}"
        expr = f"{start}:{end}"

        if ws != current_worksheet:
            expr = f"'{ws.title}'!" + expr
        return expr

    def copy(self) -> Self:
        return Range(self.start, self.end, self.range_worksheet)  # type: ignore

    def __str__(self) -> str:
        if self.worksheet is not None:
            return f"Range({self.start}, {self.end}, {self.worksheet.title})"
        else:
            return f"Range({self.start}, {self.end})"


def evaluate_expression(expr: Expression | int | float | str, interpreter=None) -> Any:
    # Importing here avoids circular references
    from excel_interpreter.interpreter import ExcelInterpreter

    if isinstance(expr, (int, float, str)):
        return expr

    if isinstance(expr, Constant):
        return expr.value

    if isinstance(expr, Cell):
        val = expr.value
        if isinstance(val, str) and val.strip().startswith("="):
            ws = expr.main_location.worksheet
            wb = ws.parent
            assert wb is not None  # this is always true AFAIK
            if interpreter is None:
                interpreter = ExcelInterpreter(wb)
            else:
                assert (
                    isinstance(interpreter, ExcelInterpreter) and interpreter.workbook == wb
                ), "Expected ExcelInterpreter for the same workbook"

            val = interpreter.evaluate(val, ws)

        if isinstance(val, str) and val.isdigit():
            val = float(val)  # type: ignore
        # Usual Excel interpretation
        if val is None:
            val = 0
        return val

    if isinstance(expr, UnaryExpression):
        match expr.op:
            case "+":
                return evaluate_expression(expr.expr, interpreter=interpreter)
            case "-":
                return -1 * evaluate_expression(expr.expr, interpreter=interpreter)
            case _:
                raise Exception(f"Unsupported unary operator: {expr.op}")

    if isinstance(expr, BinaryExpression):
        left = evaluate_expression(expr.left, interpreter=interpreter)
        right = evaluate_expression(expr.right, interpreter=interpreter)
        match expr.op:
            case "+":
                return left + right
            case "-":
                return left - right
            case "*":
                return left * right
            case "/":
                return left / right
            case ">":
                return left > right
            case ">=":
                return left >= right
            case "<":
                return left < right
            case "<=":
                return left <= right
            case "=":
                return left == right
            case "<>":
                return left != right
            case _:
                raise Exception(f"Unsupported binary operator: {expr.op}")

    if isinstance(expr, ExcelFunction):
        # Special case for IF, to avoid evaluating both branches
        if expr.symbol == "IF":
            cond = expr.arguments[0]
            ok = expr.arguments[1]
            not_ok = expr.arguments[2]

            cond = evaluate_expression(cond, interpreter=interpreter)
            assert isinstance(cond, bool), (
                f"IF condition resolved to non-boolean: {cond}"
            )
            if cond:
                return evaluate_expression(ok, interpreter=interpreter)
            else:
                return evaluate_expression(not_ok, interpreter=interpreter)

        # Special case for SUM, where we handle Range expressions
        if expr.symbol == "SUM":
            total = 0
            for arg in expr.arguments:
                if isinstance(arg, Range):
                    start, end = arg.start_and_end_locations()
                    assert start.worksheet == end.worksheet, (
                        "Can't evaluate Range expression"
                    )
                    ws = start.worksheet
                    min_row = min(start.row, end.row)
                    min_col = min(start.col, end.col)
                    max_row = max(start.row, end.row)
                    max_col = max(start.col, end.col)
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            total += evaluate_expression(Cell(ws, row, col), interpreter=interpreter)
                else:
                    total += evaluate_expression(arg, interpreter=interpreter)
            return total

        from excel_interpreter.functions import EXCEL_FUNCTIONS

        evaluated_args = [evaluate_expression(arg, interpreter=interpreter) for arg in expr.arguments]
        fn = EXCEL_FUNCTIONS.get(expr.symbol)
        if fn is None:
            raise Exception(f"Function not implemented: {fn}")
        return fn(*evaluated_args)

    raise Exception(f"Unsupported expression: {expr}")


def register_expression(wb: Workbook, expr: Expression):
    if not hasattr(wb, registration_key):
        setattr(wb, registration_key, {})
    registrar: dict[int, Expression] = getattr(wb, registration_key)
    if expr.id not in registrar:
        registrar[expr.id] = expr


simplification_cache: dict[int, ExprOrValue] | None = None


@contextmanager
def new_simplification_cache():
    global simplification_cache
    old_cache = simplification_cache
    simplification_cache = {}
    try:
        yield
    finally:
        simplification_cache = old_cache


def simplify_workbook(wb: Workbook):
    global fast_mode
    fast_mode_prev = fast_mode
    fast_mode = False
    registrar: dict[int, Expression] = getattr(wb, registration_key, {})

    if len(registrar) == 0:
        return

    visited: set[int] = set()

    def visit(expr: Expression):
        if expr.id in visited:
            return

        for dep in expr.deps():
            if isinstance(dep, Expression):
                visit(dep)

        visited.add(expr.id)
        for i in range(len(expr.locations)):
            expr.update_location(i)

    with new_simplification_cache():
        for expr in registrar.values():
            visit(expr)

    fast_mode = fast_mode_prev
