import pytest
from openpyxl import Workbook
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import quote_sheetname, absolute_coordinate
from excel_interpreter.errors import (
    CoercionError,
    CycleError,
    ExcelFunctionError,
    NameNotFound,
    SheetNotFound,
    TokenizerError,
)
from excel_interpreter.functions import flatten_args
from excel_interpreter.interpreter import ExcelInterpreter
import math
from datetime import datetime, time, timedelta


@pytest.fixture
def workbook():
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = "Sheet1"

    # Setup Sheet1 with test data
    for i in range(1, 6):
        sheet1[f"A{i}"] = i  # A1 to A5: 1, 2, 3, 4, 5
        sheet1[f"B{i}"] = i * 2  # B1 to B5: 2, 4, 6, 8, 10

    # Create Sheet2 with more test data
    sheet2 = wb.create_sheet("Sheet2")
    for i in range(1, 4):
        sheet2[f"C{i}"] = i * 3  # C1 to C3: 3, 6, 9

    return wb


@pytest.fixture
def interpreter(workbook):
    return ExcelInterpreter(workbook)


class TestBasicOperations:
    def test_simple_arithmetic(self, interpreter):
        assert interpreter.evaluate("=1 + 2", "Sheet1") == 3
        assert interpreter.evaluate("=2 * 3", "Sheet1") == 6
        assert interpreter.evaluate("=10 / 2", "Sheet1") == 5
        assert interpreter.evaluate("=5 - 3", "Sheet1") == 2

    def test_order_of_operations(self, interpreter):
        assert interpreter.evaluate("=2 + 3 * 4", "Sheet1") == 14
        assert interpreter.evaluate("=(2 + 3) * 4", "Sheet1") == 20
        assert interpreter.evaluate("=10 - 2 * 3", "Sheet1") == 4


class TestCellReferences:
    def test_single_cell_reference(self, interpreter):
        assert interpreter.evaluate("=A1", "Sheet1") == 1
        assert interpreter.evaluate("=B1", "Sheet1") == 2

    def test_cross_sheet_reference(self, interpreter):
        assert interpreter.evaluate("=Sheet2!C1", "Sheet1") == 3
        assert interpreter.evaluate("='Sheet2'!C2", "Sheet1") == 6

    def test_absolute_references(self, interpreter):
        # Test absolute column
        assert interpreter.evaluate("=$A1", "Sheet1") == 1

        # Test absolute row
        assert interpreter.evaluate("=A$1", "Sheet1") == 1

        # Test both absolute
        assert interpreter.evaluate("=$A$1", "Sheet1") == 1

        # Test in formula
        assert interpreter.evaluate("=SUM($A$1:$A$3)", "Sheet1") == 6

    def test_mixed_absolute_relative_references(self, interpreter):
        # Test mixed references in range
        assert interpreter.evaluate("=SUM($A1:A$3)", "Sheet1") == 6  # 1 + 2 + 3

        # Test with arithmetic
        assert interpreter.evaluate("=$B$1 + A$2", "Sheet1") == 4  # 2 + 2

    def test_ranges(self, interpreter):
        assert interpreter.evaluate("=A2:A2", "Sheet1") == [2]


class TestExcelFunctions:
    def test_sum_function(self, interpreter):
        assert interpreter.evaluate("=SUM(A1:A5)", "Sheet1") == 15
        assert interpreter.evaluate("=SUM(1, 2, 3)", "Sheet1") == 6
        assert interpreter.evaluate("=SUM(A1:A3, B1:B2)", "Sheet1") == 12
        assert interpreter.evaluate("=SUM(A1:B3)", "Sheet1") == 18

    def test_average_function(self, interpreter):
        assert interpreter.evaluate("=AVERAGE(A1:A5)", "Sheet1") == 3
        assert interpreter.evaluate("=AVERAGE(1, 2, 3)", "Sheet1") == 2

    def test_max_min_functions(self, interpreter):
        assert interpreter.evaluate("=MAX(A1:A5)", "Sheet1") == 5
        assert interpreter.evaluate("=MIN(A1:A5)", "Sheet1") == 1

    def test_text_date_time_formats(self, interpreter):
        sheet = interpreter.workbook["Sheet1"]
        sheet["C1"] = datetime(2026, 2, 9, 16, 30, 0)
        sheet["D1"] = time(16, 0, 0)

        assert interpreter.evaluate('=TEXT(C1, "mmdd")', "Sheet1") == "0209"
        assert interpreter.evaluate('=TEXT(C1, "hh")', "Sheet1") == "16"
        assert interpreter.evaluate('=TEXT(D1, "hh")', "Sheet1") == "16"


class TestBooleanOperations:
    @pytest.fixture
    def bool_workbook(self):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Sheet1"
        sheet["A1"] = True
        sheet["A2"] = False
        sheet["B1"] = 1
        sheet["B2"] = 0
        return wb

    @pytest.fixture
    def bool_interpreter(self, bool_workbook):
        return ExcelInterpreter(bool_workbook)

    def test_boolean_constants(self, bool_interpreter):
        assert bool_interpreter.evaluate("=TRUE", "Sheet1") is True
        assert bool_interpreter.evaluate("=FALSE", "Sheet1") is False

    def test_boolean_operations(self, bool_interpreter):
        assert bool_interpreter.evaluate("=AND(A1, A2)", "Sheet1") is False
        assert bool_interpreter.evaluate("=OR(A1, A2)", "Sheet1") is True
        assert bool_interpreter.evaluate("=NOT(A1)", "Sheet1") is False

    def test_boolean_arithmetic(self, bool_interpreter):
        assert bool_interpreter.evaluate("=A1 + B1", "Sheet1") == 2
        assert bool_interpreter.evaluate("=A2 + B2", "Sheet1") == 0


class TestErrorHandling:
    def test_division_by_zero(self, interpreter):
        assert interpreter.evaluate("=1/0", "Sheet1") == "#DIV/0!"

    def test_invalid_cell_reference(self, interpreter):
        with pytest.raises(NameNotFound):
            interpreter.evaluate("=InvalidCell", "Sheet1")

    def test_invalid_sheet_reference(self, interpreter):
        with pytest.raises(SheetNotFound):
            interpreter.evaluate("=InvalidSheet!A1", "Sheet1")


class TestCaching:
    def test_cache_invalidation(self, interpreter):
        # First evaluation should cache the result
        result1 = interpreter.evaluate("=A1", "Sheet1")

        # Modify the cell value
        interpreter.workbook["Sheet1"]["A1"].value = 100

        # Without cache invalidation, should return old value
        result2 = interpreter.evaluate("=A1", "Sheet1")
        assert result2 == result1

        # After cache invalidation, should return new value
        interpreter.invalidate_cell("Sheet1", 1, 1)
        result3 = interpreter.evaluate("=A1", "Sheet1")
        assert result3 == 100

    def test_cache_clear(self, interpreter):
        # Cache some values
        interpreter.evaluate("=A1", "Sheet1")
        interpreter.evaluate("=B1", "Sheet1")

        # Clear cache
        interpreter.clear_cache()
        assert len(interpreter.cache) == 0


class TestRanges:
    def test_cell_range(self, interpreter):
        result = interpreter.evaluate("=A1:A3", "Sheet1")
        assert result == [1, 2, 3]

    def test_ranges_across_sheets(self, interpreter):
        assert interpreter.evaluate("=SUM(Sheet1!A1:'Sheet1'!A5)", "Sheet2") == 15


class TestNamedRanges:
    @pytest.fixture
    def workbook_with_names(self):
        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = "Sheet1"
        sheet2 = wb.create_sheet("Sheet2")

        # Setup test data
        for i in range(1, 6):
            sheet1[f"A{i}"] = i  # A1 to A5: 1, 2, 3, 4, 5
            sheet1[f"B{i}"] = i * 2  # B1 to B5: 2, 4, 6, 8, 10
        for i in range(1, 4):
            sheet2[f"C{i}"] = i * 3  # C1 to C3: 3, 6, 9

        # Define workbook-level names
        wb.defined_names.add(
            DefinedName(
                name="MyRange",  # Used for range operations (A1:A3)
                attr_text=f"{quote_sheetname('Sheet1')}!{absolute_coordinate('A1:A3')}",
            )
        )
        wb.defined_names.add(
            DefinedName(
                name="GlobalCell",  # Used for scoping tests (A1)
                attr_text=f"{quote_sheetname('Sheet1')}!{absolute_coordinate('A1')}",
            )
        )

        # Define sheet-level names
        sheet1.defined_names.add(
            DefinedName(
                name="LocalCell",  # Used for scoping tests (B1 in Sheet1)
                attr_text=f"{quote_sheetname('Sheet1')}!{absolute_coordinate('B1')}",
            )
        )

        sheet2.defined_names.add(
            DefinedName(
                name="LocalCell",  # Used for scoping tests (C1 in Sheet2)
                attr_text=f"{quote_sheetname('Sheet2')}!{absolute_coordinate('C1')}",
            )
        )

        return wb

    @pytest.fixture
    def name_interpreter(self, workbook_with_names):
        return ExcelInterpreter(workbook_with_names)

    def test_use_named_range(self, name_interpreter):
        # Use the workbook-level named range
        result = name_interpreter.evaluate("=SUM(MyRange)", "Sheet1")
        assert result == 6  # Sum of A1:A3 (1+2+3)

    def test_sheet_scoped_name(self, name_interpreter):
        # Test sheet-specific references
        assert name_interpreter.evaluate("=LocalCell", "Sheet1") == 2  # B1 in Sheet1
        assert name_interpreter.evaluate("=LocalCell", "Sheet2") == 3  # C1 in Sheet2

    def test_named_range_in_formula(self, name_interpreter):
        # Use named range in formula with cell references
        result = name_interpreter.evaluate("=SUM(MyRange) * A4", "Sheet1")
        assert result == 24  # (1+2+3) * 4

    def test_name_scoping(self, name_interpreter):
        """Test Excel name scoping rules:
        1. Sheet-scoped names are only visible in their sheet
        2. Global names are visible everywhere
        3. Sheet-scoped names take precedence over global names
        """
        # Global name should be accessible from any sheet
        assert name_interpreter.evaluate("=GlobalCell", "Sheet1") == 1
        assert name_interpreter.evaluate("=GlobalCell", "Sheet2") == 1

        # Local names should only be accessible in their sheet
        assert name_interpreter.evaluate("=LocalCell", "Sheet1") == 2  # B1 in Sheet1
        assert name_interpreter.evaluate("=LocalCell", "Sheet2") == 3  # C1 in Sheet2

        # Local names should not be accessible from other sheets
        name_interpreter.workbook.create_sheet("Sheet3")
        with pytest.raises(NameNotFound, match="Undefined name: LocalCell"):
            name_interpreter.evaluate("=LocalCell", "Sheet3")


class TestMathFunctions:
    def test_basic_math_functions(self, interpreter):
        """Test basic mathematical functions."""
        assert interpreter.evaluate("=PI()", "Sheet1") == math.pi
        assert interpreter.evaluate("=ABS(-3.14)", "Sheet1") == 3.14
        assert interpreter.evaluate("=SQRT(16)", "Sheet1") == 4
        assert interpreter.evaluate("=POWER(2, 3)", "Sheet1") == 8

    def test_trunc_function(self, interpreter):
        """Test truncation function."""
        assert interpreter.evaluate("=TRUNC(3.14159)", "Sheet1") == 3
        assert interpreter.evaluate("=TRUNC(3.14159, 2)", "Sheet1") == 3.14
        assert interpreter.evaluate("=TRUNC(-3.14159)", "Sheet1") == -3
        assert interpreter.evaluate("=TRUNC(-3.14159, 2)", "Sheet1") == -3.14

    def test_ceiling_function(self, interpreter):
        """Test ceiling function."""
        assert interpreter.evaluate("=CEILING(2.5)", "Sheet1") == 3
        assert interpreter.evaluate("=CEILING(2.5, 0.5)", "Sheet1") == 2.5
        assert interpreter.evaluate("=CEILING(-2.5)", "Sheet1") == -2
        assert interpreter.evaluate("=CEILING(0)", "Sheet1") == 0

        # Test error cases
        with pytest.raises(ExcelFunctionError):
            interpreter.evaluate("=CEILING(2.5, 0)", "Sheet1")
        with pytest.raises(ExcelFunctionError):
            interpreter.evaluate("=CEILING(2.5, -1)", "Sheet1")

    def test_exponential_functions(self, interpreter):
        """Test exponential and logarithmic functions."""
        assert interpreter.evaluate("=EXP(0)", "Sheet1") == 1
        assert interpreter.evaluate("=EXP(1)", "Sheet1") == math.e
        assert abs(interpreter.evaluate("=LN(2.718281828459045)", "Sheet1") - 1) < 1e-10

        # Test error cases
        with pytest.raises(ExcelFunctionError):
            interpreter.evaluate("=LN(0)", "Sheet1")
        with pytest.raises(ExcelFunctionError):
            interpreter.evaluate("=LN(-1)", "Sheet1")

    def test_npv_function(self, interpreter):
        """Test Net Present Value calculations."""
        # Test basic NPV calculation
        assert (
            abs(
                interpreter.evaluate("=NPV(0.1, 100, 200)", "Sheet1")
                - (100 / 1.1 + 200 / 1.21)
            )
            < 0.01
        )

        # Test with cell ranges
        interpreter.workbook["Sheet1"]["D1"] = 0.1  # rate
        interpreter.workbook["Sheet1"]["D2"] = 100  # value 1
        interpreter.workbook["Sheet1"]["D3"] = 200  # value 2
        assert (
            abs(
                interpreter.evaluate("=NPV(D1, D2:D3)", "Sheet1")
                - (100 / 1.1 + 200 / 1.21)
            )
            < 0.01
        )

        # Test error cases
        with pytest.raises(ExcelFunctionError):
            interpreter.evaluate("=NPV(0.1)", "Sheet1")  # Not enough arguments


class TestEmptyCells:
    @pytest.fixture
    def empty_workbook(self):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Sheet1"

        # Setup mixed data with empty cells
        sheet["A1"] = 1
        sheet["A2"] = None  # Empty cell
        sheet["A3"] = ""  # Empty string
        sheet["A4"] = 4

        sheet["B1"] = "Text"
        sheet["B2"] = None  # Empty cell
        sheet["B3"] = ""  # Empty string
        sheet["B4"] = "More"

        return wb

    @pytest.fixture
    def empty_interpreter(self, empty_workbook):
        return ExcelInterpreter(empty_workbook)

    def test_empty_cells_in_arithmetic(self, empty_interpreter):
        """Test how empty cells are handled in arithmetic operations."""
        # Empty cells should be treated as 0 in arithmetic
        assert empty_interpreter.evaluate("=A1 + A2", "Sheet1") == 1  # 1 + (empty) = 1
        # A3 is a string so it should fail on arithmetic operations
        with pytest.raises(CoercionError):
            assert empty_interpreter.evaluate("=A2 + A3", "Sheet1") == 0
        assert empty_interpreter.evaluate("=SUM(A1:A4)", "Sheet1") == 5  # 1 + 0 + 0 + 4

    def test_empty_cells_in_functions(self, empty_interpreter):
        """Test how empty cells are handled in various functions."""
        # AVERAGE should ignore empty cells
        assert (
            empty_interpreter.evaluate("=AVERAGE(A1:A4)", "Sheet1") == 2.5
        )  # (1 + 4) / 2

        # MAX/MIN should ignore empty cells
        assert empty_interpreter.evaluate("=MAX(A1:A4)", "Sheet1") == 4
        assert empty_interpreter.evaluate("=MIN(A1:A4)", "Sheet1") == 1

    def test_empty_cells_in_comparisons(self, empty_interpreter):
        """Test how empty cells are handled in comparisons."""
        # Empty cells should be equal to each other
        assert empty_interpreter.evaluate("=Z1 = Z2", "Sheet1") == True
        # Empty cells should be equal to zero in numeric contexts
        assert empty_interpreter.evaluate("=Z1 = 0", "Sheet1") == True
        # Empty cells should not be equal to empty string in text contexts
        assert empty_interpreter.evaluate('=Z1 = ""', "Sheet1") == False


class TestTypeCoercion:
    """Test Excel's type coercion rules."""

    def test_numeric_coercion(self, interpreter):
        """Test coercion to numbers."""

        def evaluate_case(interpreter, formula, expected):
            if isinstance(expected, type) and issubclass(expected, Exception):
                with pytest.raises(expected):
                    interpreter.evaluate(formula, "Sheet1")
            else:
                if isinstance(expected, float):
                    result = interpreter.evaluate(formula, "Sheet1")
                    assert abs(result - expected) < 1e-10
                else:
                    assert interpreter.evaluate(formula, "Sheet1") == expected

        # Simple numbers
        evaluate_case(interpreter, "=1+1", 2)
        # String to number
        evaluate_case(interpreter, '="123"+1', 124)
        # Boolean to number (TRUE = 1)
        evaluate_case(interpreter, "=TRUE+1", 2)
        # Boolean to number (FALSE = 0)
        evaluate_case(interpreter, "=FALSE+1", 1)
        # Empty string does not convert
        evaluate_case(interpreter, '=""+1', CoercionError)
        # Empty cell to 0
        evaluate_case(interpreter, "=Z1+1", 1)
        # Invalid string to number
        evaluate_case(interpreter, '=1+"abc"', CoercionError)
        # General string addition should fail
        evaluate_case(interpreter, '="abc"+"def"', CoercionError)
        # Except for numeric strings
        evaluate_case(interpreter, '="123"+"456"', 579)

        # Scientific notation tests
        evaluate_case(interpreter, "=1.23e5", 123000.0)
        evaluate_case(interpreter, "=4.56E+3", 4560.0)
        evaluate_case(interpreter, "=7.89e-4", 0.000789)
        evaluate_case(interpreter, "=1E-10", 1e-10)
        evaluate_case(interpreter, "=2.5e+5", 250000.0)

        # Test scientific notation in strings
        evaluate_case(interpreter, '=1+"23e-2"', 1.23)
        evaluate_case(interpreter, '=1+"23E-2"', 1.23)

        # Test scientific notation in calculations
        evaluate_case(interpreter, "=1.5e3 + 2.4e2", 1740.0)
        evaluate_case(interpreter, "=3e3 * 2e-3", 6.0)
        evaluate_case(interpreter, "=1e6 / 1e3", 1000.0)

    def test_boolean_coercion(self, interpreter):
        """Test coercion to boolean."""
        assert interpreter.evaluate("=IF(TRUE, 1, 0)", "Sheet1") == 1
        assert interpreter.evaluate("=IF(FALSE, 1, 0)", "Sheet1") == 0
        assert (
            interpreter.evaluate("=IF(1, 1, 0)", "Sheet1") == 1
        )  # Non-zero number is True
        assert interpreter.evaluate("=IF(0, 1, 0)", "Sheet1") == 0  # Zero is False
        assert (
            interpreter.evaluate('=IF("TRUE", 1, 0)', "Sheet1") == 1
        )  # Case-insensitive "TRUE"
        assert (
            interpreter.evaluate('=IF("false", 1, 0)', "Sheet1") == 0
        )  # Case-insensitive "FALSE"
        # Empty string is not coerced
        with pytest.raises(CoercionError):
            assert interpreter.evaluate('=IF("", 1, 0)', "Sheet1") == 0
        assert (
            interpreter.evaluate("=IF(Z1, 1, 0)", "Sheet1") == 0
        )  # Empty cell is False

    def test_text_coercion(self, interpreter):
        """Test coercion to text."""
        assert (
            interpreter.evaluate('="a"&"b"', "Sheet1") == "ab"
        )  # String concatenation
        assert interpreter.evaluate("=1&2", "Sheet1") == "12"  # Numbers to string
        assert (
            interpreter.evaluate('=TRUE&"x"', "Sheet1") == "TRUEx"
        )  # Boolean to "TRUE"/"FALSE"
        assert interpreter.evaluate('=FALSE&"x"', "Sheet1") == "FALSEx"
        assert (
            interpreter.evaluate('=""&Z1', "Sheet1") == ""
        )  # Empty cell to empty string

    def test_int_float_comparisons(self, interpreter):
        # This used to fail as the comparison logic that handles type
        # differences considered the two sides to be different (`int` vs
        # `float`) and compared their respective priorities (both equal to 1),
        # thus returning 1 < 1 = False. Whereas the correct result is obviously
        # True.
        assert interpreter.evaluate("=216<219.10", "Sheet1")

    def test_array_coercion(self, interpreter):
        """Test array handling and coercion."""
        assert interpreter.evaluate("=SUM({1,2,3})", "Sheet1") == 6  # Simple array
        # Mixed types (TRUE=1)
        assert interpreter.evaluate('=SUM({1,"2",TRUE})', "Sheet1") == 4
        # Invalid text ignored
        assert interpreter.evaluate('=SUM({1,"abc",3})', "Sheet1") == 4
        # Nested arrays (note: this doesn't work in Excel)
        assert interpreter.evaluate("=SUM({1,{2,3},4})", "Sheet1") == 10
        # Empty cells
        assert interpreter.evaluate("=SUM({1,Z1,3})", "Sheet1") == 4

    def test_comparison_coercion(self, interpreter):
        """Test type coercion in comparisons."""
        # String are not equal to numbers
        assert interpreter.evaluate('="123"=123', "Sheet1") == False
        # Case-insensitive string comparison
        assert interpreter.evaluate('="abc"="ABC"', "Sheet1") == True
        # Empty string is not coerced to zero
        assert interpreter.evaluate('=""=0', "Sheet1") == False
        # Empty cell equals zero
        assert interpreter.evaluate("=Z1=0", "Sheet1") == True
        # String comparison
        assert interpreter.evaluate('="abd">"abc"', "Sheet1") == True
        # Strings are always larger than numbers
        assert interpreter.evaluate('="1">2', "Sheet1") == True
        # Boolean/number equality
        assert interpreter.evaluate("=TRUE=1", "Sheet1") == False

    def test_date_coercion(self, interpreter):
        """Test date arithmetic and coercion."""
        interpreter.workbook["Sheet1"]["A1"] = datetime(2023, 1, 1)
        interpreter.workbook["Sheet1"]["A2"] = datetime(2023, 1, 6)
        # Date + number
        assert interpreter.evaluate("=A1+1", "Sheet1") == datetime(2023, 1, 2)
        # Date - date = timedelta
        assert interpreter.evaluate("=A2-A1", "Sheet1") == timedelta(5)

    def test_error_propagation(self, interpreter):
        """Test error handling in calculations."""
        assert interpreter.evaluate("=1/0", "Sheet1") == "#DIV/0!"
        with pytest.raises(ValueError):
            interpreter.evaluate("=SQRT(-1)", "Sheet1")  # Invalid sqrt
        with pytest.raises(CoercionError):
            interpreter.evaluate('="abc"+1', "Sheet1")  # Invalid conversion

    def test_function_type_handling(self, interpreter):
        """Test how functions handle different types."""

        def check_output(formula: str, expected):
            assert interpreter.evaluate(formula, "Sheet1") == expected

        # Mixed types in SUM
        check_output('=SUM(1,"2",TRUE)', 4)
        # Mixed types in MIN
        check_output('=MIN("999",1,TRUE)', 1)
        # Mixed types in AND
        check_output('=AND(1,TRUE,"TRUE")', True)
        # Mixed types in OR
        check_output('=OR(0,FALSE,"abc")', False)
        with pytest.raises(CoercionError):
            check_output('=IF("abc",1,0)', 1)


class TestStringLiterals:
    def test_string_literals(self, interpreter):
        """Test Excel string literal handling."""
        cases = [
            ('="Simple string"', "Simple string"),
            ('="String with ""quotes"""', 'String with "quotes"'),
            ('="Multi\nline"', "Multi\nline"),
            ('="Special chars: !@#$%^&*()"', "Special chars: !@#$%^&*()"),
            ('="Mixed""quote"', 'Mixed"quote'),
            ('="Trailing quote"""', 'Trailing quote"'),
            ('=""', ""),  # Empty string
            # Raw strings (no formula)
            ("Raw string", "Raw string"),
            ('String with "quotes"', 'String with "quotes"'),
        ]
        for formula, expected in cases:
            assert interpreter.evaluate(formula, "Sheet1") == expected

    def test_string_literal_errors(self, interpreter):
        """Test string literal error handling."""
        with pytest.raises(TokenizerError, match="Unterminated string literal"):
            interpreter.evaluate('="Missing end quote', "Sheet1")


class TestQuotedIdentifiers:
    def test_quoted_identifiers(self, interpreter):
        """Test handling of quoted identifiers."""

        # Setup test data
        wb = Workbook()
        wb.active.title = "Sheet1"
        interpreter = ExcelInterpreter(wb)
        for sheet_name in ["Sheet 1", "My-Sheet", "Bob's Sheet", "Sheet 2"]:
            if sheet_name not in wb:
                wb.create_sheet(sheet_name)

        wb["Sheet 1"]["A1"] = "value1"
        wb["My-Sheet"]["B2"] = "value2"
        wb["Bob's Sheet"]["C3"] = "value3"
        wb["Sheet 2"]["B2"] = "value4"

        # Sheet names with spaces
        assert interpreter.evaluate("='Sheet 1'!A1", "Sheet1") == "value1"
        # Sheet names with special characters
        assert interpreter.evaluate("='My-Sheet'!B2", "Sheet1") == "value2"
        # Escaped single quotes in sheet names
        assert interpreter.evaluate("='Bob''s Sheet'!C3", "Sheet1") == "value3"
        # Multiple parts
        assert (
            interpreter.evaluate("='Sheet 1'!A1 & 'Sheet 2'!B2", "Sheet1")
            == "value1value4"
        )
        # Mixed with regular identifiers
        assert interpreter.evaluate("=A1&'Sheet 2'!B2", "Sheet1") == "value4"

    def test_quoted_identifier_errors(self, interpreter):
        """Test error handling for quoted identifiers."""
        with pytest.raises(TokenizerError, match="Unterminated quoted identifier"):
            interpreter.evaluate("='Unterminated", "Sheet1")

        with pytest.raises(SheetNotFound, match='Worksheet "Invalid Sheet" not found'):
            interpreter.evaluate("='Invalid Sheet'!A1", "Sheet1")


class TestFlattenArgs:
    """Test the flatten_args function that flattens Excel function arguments."""

    def test_flat_args(self):
        """Test flattening of non-array arguments."""
        args = [1, "text", True]
        result = flatten_args(*args)
        assert result == args

    def test_1d_array(self):
        """Test flattening of 1D array arguments."""
        args = [1, [2, 3], 4]
        result = flatten_args(*args)
        assert result == [1, 2, 3, 4]

    def test_2d_array(self):
        """Test flattening of 2D array arguments."""
        args = [1, [[2, 3], [4, 5]], 6]
        result = flatten_args(*args)
        assert result == [1, 2, 3, 4, 5, 6]

    def test_mixed_arrays(self):
        """Test flattening of mixed 1D and 2D arrays with different types."""
        args = [[1, True], [["a", "b"], [2, False]], "text"]
        result = flatten_args(*args)
        assert result == [1, True, "a", "b", 2, False, "text"]


class TestOperatorPrecedence:
    def test_arithmetic_precedence(self, interpreter):
        """Test arithmetic operator precedence."""
        # Basic arithmetic precedence
        assert (
            interpreter.evaluate("=2 + 3 * 4", "Sheet1") == 14
        )  # Multiplication before addition
        assert (
            interpreter.evaluate("=2 * 3 + 4", "Sheet1") == 10
        )  # Multiplication before addition
        assert (
            interpreter.evaluate("=10 - 2 * 3", "Sheet1") == 4
        )  # Multiplication before subtraction

        # Exponentiation precedence (highest)
        assert (
            interpreter.evaluate("=2 ^ 3 * 2", "Sheet1") == 16
        )  # Exponentiation before multiplication
        assert (
            interpreter.evaluate("=2 * 3 ^ 2", "Sheet1") == 18
        )  # Exponentiation before multiplication
        assert (
            interpreter.evaluate("=2 + 2 ^ 3", "Sheet1") == 10
        )  # Exponentiation before addition
        assert (
            interpreter.evaluate("=2 ^ 2 ^ 3", "Sheet1") == 64
        )  # Right-associative: 2^(2^3)

        # Complex arithmetic expressions
        assert interpreter.evaluate("=2 ^ 3 * 4 + 5", "Sheet1") == 37  # 8 * 4 + 5
        assert interpreter.evaluate("=10 - 2 ^ 3 + 4", "Sheet1") == 6  # 10 - 8 + 4

    def test_string_concatenation_precedence(self, interpreter):
        """Test string concatenation precedence (lowest)."""
        # Concatenation with arithmetic
        assert (
            interpreter.evaluate('=2 + 3 & "x"', "Sheet1") == "5x"
        )  # Addition before concatenation
        assert (
            interpreter.evaluate('="x" & 2 + 3', "Sheet1") == "x5"
        )  # Addition before concatenation
        assert (
            interpreter.evaluate('=2 * 3 & "x"', "Sheet1") == "6x"
        )  # Multiplication before concatenation

        # Concatenation with exponentiation
        assert (
            interpreter.evaluate('=2 ^ 3 & "x"', "Sheet1") == "8x"
        )  # Exponentiation before concatenation
        assert (
            interpreter.evaluate('="x" & 2 ^ 3', "Sheet1") == "x8"
        )  # Exponentiation before concatenation

        # Multiple concatenations
        assert (
            interpreter.evaluate('="a" & "b" & "c"', "Sheet1") == "abc"
        )  # Left-associative
        assert (
            interpreter.evaluate('="a" & 2 + 3 & "c"', "Sheet1") == "a5c"
        )  # Arithmetic before concatenation

    def test_comparison_precedence(self, interpreter):
        """Test comparison operator precedence."""
        # Arithmetic before comparison
        assert interpreter.evaluate("=2 + 3 > 4", "Sheet1") is True  # 5 > 4
        assert interpreter.evaluate("=2 * 3 = 6", "Sheet1") is True  # 6 = 6

        # Exponentiation before comparison
        assert interpreter.evaluate("=2 ^ 3 > 7", "Sheet1") is True  # 8 > 7
        assert interpreter.evaluate("=2 ^ 3 = 8", "Sheet1") is True  # 8 = 8

        # Concatenation before comparison
        assert interpreter.evaluate('=2 > 1 & "true"', "Sheet1") == False
        assert interpreter.evaluate('="22" = 2 & 2', "Sheet1") == True

    def test_complex_precedence(self, interpreter):
        """Test complex expressions with multiple operators."""
        # Complex arithmetic with concatenation
        assert interpreter.evaluate('=2 ^ 3 * 4 + 5 & "x"', "Sheet1") == "37x"

        # Mixed arithmetic and comparison
        assert interpreter.evaluate("=2 ^ 3 * 4 > 30", "Sheet1") is True

        # Everything together
        assert (
            interpreter.evaluate('=2 ^ 2 * 3 + 1 > 10 & "!" & 2 ^ 3', "Sheet1") == False
        )

        # Parentheses overriding precedence
        assert interpreter.evaluate("=(2 + 3) * 4", "Sheet1") == 20
        assert interpreter.evaluate("=2 ^ (3 + 1)", "Sheet1") == 16
        assert interpreter.evaluate('=(2 + 3 > 4) & "test"', "Sheet1") == "TRUEtest"


class TestDateTimeOperations:
    @pytest.fixture
    def date_workbook(self):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Sheet1"

        # Setup test data with dates and times
        sheet["A1"] = datetime(2023, 1, 1)  # New Year's Day
        sheet["A2"] = datetime(2023, 1, 6)  # 5 days later
        sheet["A3"] = datetime(2023, 1, 1, 12, 30)  # Same day with time
        sheet["A4"] = datetime(2023, 12, 31, 23, 59)  # Year end
        sheet["B1"] = 5  # Days to add
        sheet["B2"] = 1.5  # Days and hours to add

        # Add timedelta test data
        sheet["C1"] = timedelta(days=2, hours=12)  # 2.5 days
        sheet["C2"] = timedelta(hours=36)  # 1.5 days
        sheet["C3"] = timedelta(minutes=90)  # 1.5 hours
        sheet["C4"] = timedelta(days=1, hours=12, minutes=30)  # 1.52083... days
        sheet["D1"] = 2.5  # Matches C1's days for multiplication tests
        return wb

    @pytest.fixture
    def date_interpreter(self, date_workbook):
        return ExcelInterpreter(date_workbook)

    def test_date_to_number_coercion(self, date_interpreter):
        """Test date to number coercion including Excel's 1900 leap year quirk."""
        # Basic date to number coercion
        assert date_interpreter.evaluate("=A1*1", "Sheet1") == 44927  # 2023-01-01

        # Test dates around the 1900 leap year quirk
        assert date_interpreter.evaluate("=DATE(1900,2,28)*1", "Sheet1") == 59
        # Skips day 60, 1900/2/29
        assert date_interpreter.evaluate("=DATE(1900,3,1)*1", "Sheet1") == 61
        assert date_interpreter.evaluate("=DATE(1900,3,1)=61", "Sheet1") == True
        assert date_interpreter.evaluate("=DATE(1900,3,2)*1", "Sheet1") == 62

        # Test early dates
        assert date_interpreter.evaluate("=DATE(1900,1,1)*1", "Sheet1") == 1
        assert date_interpreter.evaluate("=DATE(1900,1,2)*1", "Sheet1") == 2

        # Test modern dates
        assert date_interpreter.evaluate("=DATE(2000,1,1)*1", "Sheet1") == 36526

        # Test datetimes

    def test_date_arithmetic(self, date_interpreter):
        """Test basic date arithmetic operations."""
        # Date + number (days)
        assert date_interpreter.evaluate("=A1+1", "Sheet1") == datetime(2023, 1, 2)
        assert date_interpreter.evaluate("=A1+B1", "Sheet1") == datetime(2023, 1, 6)

        # Date - number (days)
        assert date_interpreter.evaluate("=A2-3", "Sheet1") == datetime(2023, 1, 3)

        # Date - date = days
        assert date_interpreter.evaluate("=A2-A1", "Sheet1") == timedelta(5)

        # Date + fractional days
        expected_time = datetime(2023, 1, 2, 12, 0)  # 1.5 days = 36 hours
        result = date_interpreter.evaluate("=A1+B2", "Sheet1")
        assert (
            abs((result - expected_time).total_seconds()) < 1
        )  # Allow 1 second difference

        assert date_interpreter.evaluate(
            '=TIME(12,30,0)+"01:00"', "Sheet1"
        ) == timedelta(hours=13, minutes=30)

        assert date_interpreter.evaluate("=DATE(1900,2,28) + 2", "Sheet1") == datetime(
            1900, 3, 1, 0, 0, 0
        )

    def test_time_arithmetic(self, date_interpreter):
        """Test time-specific arithmetic operations."""
        # Time difference within same day
        time_diff = date_interpreter.evaluate("=A3-A1", "Sheet1")
        assert time_diff == timedelta(seconds=12.5 * 3600)

        # Add hours using decimal days
        result = date_interpreter.evaluate("=A1+1/24", "Sheet1")  # Add 1 hour
        assert result == datetime(2023, 1, 1, 1, 0)

        # Add minutes using decimal days
        result = date_interpreter.evaluate("=A1+1/1440", "Sheet1")  # Add 1 minute
        assert result == datetime(2023, 1, 1, 0, 1)


    def test_date_comparisons(self, date_interpreter):
        """Test date comparison operations."""
        assert date_interpreter.evaluate("=A2>A1", "Sheet1") is True
        assert date_interpreter.evaluate("=A1<A2", "Sheet1") is True
        assert date_interpreter.evaluate("=A1=A1", "Sheet1") is True
        assert date_interpreter.evaluate("=A1>=A1", "Sheet1") is True
        assert date_interpreter.evaluate("=A2<=A2", "Sheet1") is True
        assert date_interpreter.evaluate("=A1<>A2", "Sheet1") is True

    def test_date_functions(self, date_interpreter):
        """Test date-specific functions."""
        # YEAR function
        assert date_interpreter.evaluate("=YEAR(A1)", "Sheet1") == 2023

        # MONTH function
        assert date_interpreter.evaluate("=MONTH(A1)", "Sheet1") == 1

        # DAY function
        assert date_interpreter.evaluate("=DAY(A1)", "Sheet1") == 1

        # HOUR function
        assert date_interpreter.evaluate("=HOUR(A3)", "Sheet1") == 12

        # MINUTE function
        assert date_interpreter.evaluate("=MINUTE(A3)", "Sheet1") == 30

    def test_date_string_coercion(self, date_interpreter):
        """Test edge cases and special scenarios with dates."""
        # Year transition

        assert date_interpreter.evaluate('="2023-01-01" + C1', "Sheet1") == datetime(
            2023, 1, 3, 12
        )
        assert date_interpreter.evaluate('=A1 + "12:05:36"', "Sheet1") == datetime(
            2023, 1, 1, 12, 5, 36
        )
        assert date_interpreter.evaluate('=HOUR("12:30:00")', "Sheet1") == 12

    def test_timedelta_arithmetic(self, date_interpreter):
        """Test arithmetic operations with timedelta objects."""
        # Timedelta + Timedelta
        assert date_interpreter.evaluate("=C1+C2", "Sheet1") == timedelta(days=4)

        # Timedelta * Number
        assert date_interpreter.evaluate("=C1*D1", "Sheet1") == timedelta(days=6.25)
        assert date_interpreter.evaluate("=D1*C1", "Sheet1") == timedelta(days=6.25)

        # Timedelta / Number
        assert date_interpreter.evaluate("=C1/2", "Sheet1") == timedelta(days=1.25)

        # Timedelta / Timedelta (ratio)
        assert date_interpreter.evaluate("=C1/C2", "Sheet1") == 5 / 3

        # Mixed operations
        assert date_interpreter.evaluate("=(C1+C2)*2", "Sheet1") == timedelta(days=8)
        assert date_interpreter.evaluate("=C1+C3", "Sheet1") == timedelta(days=2.5625)

        # Test with date addition
        assert date_interpreter.evaluate("=A1+C3", "Sheet1") == datetime(
            2023, 1, 1, 1, 30
        )
        assert date_interpreter.evaluate("=A1+C1", "Sheet1") == datetime(2023, 1, 3, 12)

    def test_timedelta_comparisons(self, date_interpreter):
        """Test comparison operations with timedelta objects."""
        assert date_interpreter.evaluate("=C1>C2", "Sheet1") is True  # 2.5 > 1.5 days
        assert (
            date_interpreter.evaluate("=C2>C3", "Sheet1") is True
        )  # 1.5 days > 1.5 hours
        assert date_interpreter.evaluate("=C1=2.5", "Sheet1") is True  # 2.5 days == 2.5
        assert (
            date_interpreter.evaluate("=C3<C2", "Sheet1") is True
        )  # 1.5 hours < 1.5 days

        # Complex comparisons
        assert (
            date_interpreter.evaluate("=C1+C2>C4", "Sheet1") is True
        )  # 4 days > 1.52 days
        assert (
            date_interpreter.evaluate("=C1*2=5", "Sheet1") is True
        )  # 2.5 * 2 == 5 days

    def test_timedelta_edge_cases(self, date_interpreter):
        """Test edge cases and special scenarios with timedelta objects."""
        # Zero timedelta
        assert date_interpreter.evaluate("=C1-C1", "Sheet1") == timedelta(0)

        # Negative timedelta
        assert date_interpreter.evaluate("=C2-C1", "Sheet1") == timedelta(
            -1.0
        )  # 1.5 - 2.5 days

        # Very small timedelta
        small_td = timedelta(microseconds=1)
        date_interpreter.workbook["Sheet1"]["D2"] = small_td
        assert date_interpreter.evaluate("=D2", "Sheet1") == timedelta(microseconds=1)

    def test_timedelta_functions(self, date_interpreter):
        """Test Excel functions with timedelta values."""
        # SUM with timedeltas
        assert (
            date_interpreter.evaluate("=SUM(C1:C2)", "Sheet1") == 4.0
        )  # 2.5 + 1.5 days

        # MIN/MAX with timedeltas
        assert (
            date_interpreter.evaluate("=MIN(C1:C4)", "Sheet1") == 0.0625
        )  # C3 (1.5 hours)
        assert (
            date_interpreter.evaluate("=MAX(C1:C4)", "Sheet1") == 2.5
        )  # C1 (2.5 days)

        # AVERAGE with timedeltas
        avg = date_interpreter.evaluate("=AVERAGE(C1:C4)", "Sheet1")
        expected_avg = (2.5 + 1.5 + 0.0625 + 1.52083) / 4
        assert abs(avg - expected_avg) < 0.0001


class TestArrayLiterals:
    def test_horizontal_array(self, interpreter):
        """Test horizontal array literals."""
        result = interpreter.evaluate("={1,2,3}", "Sheet1")
        assert result == [1, 2, 3]

    def test_vertical_array(self, interpreter):
        """Test vertical array literals."""
        result = interpreter.evaluate("={1;2;3}", "Sheet1")
        expected = [[1], [2], [3]]
        assert result == expected

    def test_mixed_types_array(self, interpreter):
        """Test arrays with mixed types."""
        result = interpreter.evaluate('={1,"text",TRUE}', "Sheet1")
        assert result == [1, "text", True]

    def test_vertical_mixed_types(self, interpreter):
        """Test vertical arrays with mixed types."""
        result = interpreter.evaluate('={1;"text";TRUE}', "Sheet1")
        expected = [[1], ["text"], [True]]
        assert result == expected

    def test_array_in_function(self, interpreter):
        """Test using arrays in functions."""
        assert interpreter.evaluate("=SUM({1,2,3})", "Sheet1") == 6
        assert interpreter.evaluate("=SUM({1;2;3})", "Sheet1") == 6  # Vertical array
        assert interpreter.evaluate("=AVERAGE({1,2,3})", "Sheet1") == 2
        assert (
            interpreter.evaluate("=AVERAGE({1;2;3})", "Sheet1") == 2
        )  # Vertical array


class TestFormulaChaining:
    @pytest.fixture
    def chain_workbook(self):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Sheet1"

        # Basic chaining
        sheet["A1"] = "=1+1"  # 2
        sheet["A2"] = "=A1*2"  # 4
        sheet["A3"] = "=A2+A1"  # 6

        # Complex chaining with functions
        sheet["B1"] = "=10"  # 10
        sheet["B2"] = "=B1/2"  # 5
        sheet["B3"] = "=SUM(B1:B2)"  # 15
        sheet["B4"] = "=AVERAGE(B1:B3)"  # 10

        # Cross-sheet chaining
        sheet2 = wb.create_sheet("Sheet2")
        sheet2["C1"] = "=Sheet1!B4*2"  # 20
        sheet2["C2"] = "=C1+Sheet1!B3"  # 35

        # Circular chaining (should raise error)
        sheet["D1"] = "=D2"
        sheet["D2"] = "=D1"

        return wb

    @pytest.fixture
    def chain_interpreter(self, chain_workbook):
        return ExcelInterpreter(chain_workbook)

    def test_basic_chaining(self, chain_interpreter):
        """Test basic formula chaining."""
        assert chain_interpreter.evaluate("=A1", "Sheet1") == 2
        assert chain_interpreter.evaluate("=A2", "Sheet1") == 4
        assert chain_interpreter.evaluate("=A3", "Sheet1") == 6

    def test_function_chaining(self, chain_interpreter):
        """Test chaining with Excel functions."""
        assert chain_interpreter.evaluate("=B1", "Sheet1") == 10
        assert chain_interpreter.evaluate("=B2", "Sheet1") == 5
        assert chain_interpreter.evaluate("=B3", "Sheet1") == 15
        assert chain_interpreter.evaluate("=B4", "Sheet1") == 10

    def test_cross_sheet_chaining(self, chain_interpreter):
        """Test formula chains across different sheets."""
        assert chain_interpreter.evaluate("=C1", "Sheet2") == 20
        assert chain_interpreter.evaluate("=C2", "Sheet2") == 35

    def test_circular_reference(self, chain_interpreter):
        """Test that circular references raise an error."""
        with pytest.raises(CycleError):
            chain_interpreter.evaluate("=D1", "Sheet1")

    def test_deep_chaining(self, chain_interpreter):
        """Test deeply nested formula chains."""
        wb = chain_interpreter.workbook
        sheet = wb["Sheet1"]

        # Create a deep chain: each cell adds 1 to the previous
        for i in range(1, 101):  # Test with 100 levels deep
            sheet[f"E{i}"] = f"={'1' if i == 1 else f'E{i - 1}+1'}"

        # The 100th cell should equal 100
        assert chain_interpreter.evaluate("=E100", "Sheet1") == 100

    def test_complex_chaining(self, chain_interpreter):
        """Test complex chains with mixed operations and functions."""
        wb = chain_interpreter.workbook
        sheet = wb["Sheet1"]

        sheet["F1"] = "=10"
        sheet["F2"] = "=F1*2"  # 20
        sheet["F3"] = "=SUM(F1:F2)/2"  # 15
        sheet["F4"] = "=AVERAGE(F1:F3)*F1"  # (10+20+15)/3 * 10 = 150

        assert chain_interpreter.evaluate("=F4", "Sheet1") == 150


class TestCycleDetection:
    @pytest.fixture
    def cycle_workbook(self):
        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = "Sheet1"
        sheet2 = wb.create_sheet("Sheet2")

        # Setup test data for cycle detection
        # Simple cycle in Sheet1
        sheet1["A1"] = "=A2"
        sheet1["A2"] = "=A1"

        # Cross-sheet references (not a cycle)
        sheet1["B1"] = "=Sheet2!B1"
        sheet2["B1"] = "=10"

        # Same cell reference in different sheets (not a cycle)
        sheet1["C1"] = "=Sheet2!C1"
        sheet2["C1"] = "=Sheet1!C2"
        sheet1["C2"] = "=5"

        # Cross-sheet cycle
        sheet1["D1"] = "=Sheet2!D1"
        sheet2["D1"] = "=Sheet1!D2"
        sheet1["D2"] = "=Sheet1!D1"

        # Named reference cycle
        sheet1.defined_names.add(
            DefinedName(
                name="CycleName1",
                attr_text=f"{quote_sheetname('Sheet1')}!{absolute_coordinate('E2')}",
            )
        )
        sheet1["E1"] = "=CycleName1"
        sheet1["E2"] = "=E1"

        return wb

    @pytest.fixture
    def cycle_interpreter(self, cycle_workbook):
        return ExcelInterpreter(cycle_workbook)

    def test_simple_cycle(self, cycle_interpreter):
        """Test detection of a simple cycle within the same sheet."""
        with pytest.raises(CycleError) as excinfo:
            cycle_interpreter.evaluate("=A1", "Sheet1")
        assert "Detected cycle:" in str(excinfo.value)
        assert "Sheet1!A1 -> Sheet1!A2 -> Sheet1!A1" in str(excinfo.value)

    def test_cross_sheet_reference(self, cycle_interpreter):
        """Test that cross-sheet references work correctly (not a cycle)."""
        # B1 in Sheet1 references B1 in Sheet2, which is 10
        assert cycle_interpreter.evaluate("=B1", "Sheet1") == 10

    def test_same_cell_different_sheets(self, cycle_interpreter):
        """Test that same cell reference in different sheets is not detected as a cycle."""
        # C1 in Sheet1 references C1 in Sheet2, which references C2 in Sheet1, which is 5
        assert cycle_interpreter.evaluate("=C1", "Sheet1") == 5

    def test_cross_sheet_cycle(self, cycle_interpreter):
        """Test detection of a cycle that spans multiple sheets."""
        with pytest.raises(CycleError) as excinfo:
            cycle_interpreter.evaluate("=D1", "Sheet1")
        assert "Detected cycle:" in str(excinfo.value)
        # The cycle path should include sheet names
        assert "Sheet1!D1 -> Sheet2!D1 -> Sheet1!D2 -> Sheet1!D1" in str(excinfo.value)

    def test_named_reference_cycle(self, cycle_interpreter):
        """Test detection of a cycle involving named references."""
        with pytest.raises(CycleError) as excinfo:
            cycle_interpreter.evaluate("=E1", "Sheet1")
        assert "Detected cycle:" in str(excinfo.value)
        assert "name:CycleName1" in str(excinfo.value)
        assert "(in Sheet1)" in str(excinfo.value)


def test_array_formulae():
    wb = Workbook()
    ws = wb.active
    ws.title = "S"

    for i in range(1, 4):
        ws[f"A{i}"] = i
    ws["B1"] = ArrayFormula("B1:B3", "=A1:A3 + 1")
    ws["C1"] = ArrayFormula("C1:C2", "=A1:B1 * 2")
    ws["D1"] = ArrayFormula("D1:D2", "=A3:B3 * 2")

    interpreter = ExcelInterpreter(wb)
    assert interpreter.evaluate("=A1:A3", "S") == [1, 2, 3]
    for i in range(1, 4):
        assert interpreter.evaluate(f"=B{i}", "S") == i + 1

    assert interpreter.evaluate("=C1", "S") == 2  # A1 * 2
    assert interpreter.evaluate("=C2", "S") == 4  # B1 * 2
    assert interpreter.evaluate("=D1", "S") == 6  # A3 * 2
    assert interpreter.evaluate("=D2", "S") == 8  # B3 * 2


def test_countif_exact_match(interpreter):
    """Test COUNTIF with exact match criteria."""
    # Setup test data
    sheet = interpreter.workbook.active
    for i in range(1, 6):
        sheet[f"A{i}"] = i  # A1 to A5: 1, 2, 3, 4, 5

    sheet["B1"] = "apple"
    sheet["B2"] = "banana"
    sheet["B3"] = "apple"
    sheet["B4"] = "cherry"
    sheet["B5"] = "apple"

    # Test numeric exact match
    assert interpreter.evaluate("=COUNTIF(A1:A5, 3)", "Sheet1") == 1

    # Test text exact match
    assert interpreter.evaluate('=COUNTIF(B1:B5, "apple")', "Sheet1") == 3

    # Test case-insensitive match
    assert interpreter.evaluate('=COUNTIF(B1:B5, "APPLE")', "Sheet1") == 3


def test_countif_comparison_operators(interpreter):
    """Test COUNTIF with comparison operators."""
    # Setup test data
    sheet = interpreter.workbook.active
    for i in range(1, 6):
        sheet[f"A{i}"] = i  # A1 to A5: 1, 2, 3, 4, 5

    # Greater than
    assert interpreter.evaluate('=COUNTIF(A1:A5, ">2")', "Sheet1") == 3

    # Less than
    assert interpreter.evaluate('=COUNTIF(A1:A5, "<3")', "Sheet1") == 2

    # Greater than or equal
    assert interpreter.evaluate('=COUNTIF(A1:A5, ">=3")', "Sheet1") == 3

    # Less than or equal
    assert interpreter.evaluate('=COUNTIF(A1:A5, "<=3")', "Sheet1") == 3

    # Equal to
    assert interpreter.evaluate('=COUNTIF(A1:A5, "=3")', "Sheet1") == 1

    # Not equal to
    assert interpreter.evaluate('=COUNTIF(A1:A5, "<>3")', "Sheet1") == 4


def test_countif_wildcards(interpreter):
    """Test COUNTIF with wildcard patterns."""
    # Setup test data
    sheet = interpreter.workbook.active
    sheet["B1"] = "apple"
    sheet["B2"] = "banana"
    sheet["B3"] = "apricot"
    sheet["B4"] = "cherry"
    sheet["B5"] = "avocado"

    # Starts with
    assert interpreter.evaluate('=COUNTIF(B1:B5, "a*")', "Sheet1") == 3
    # Ends with
    assert interpreter.evaluate('=COUNTIF(B1:B5, "*e")', "Sheet1") == 1
    # Contains
    assert interpreter.evaluate('=COUNTIF(B1:B5, "*a*")', "Sheet1") == 4
    # Single character wildcard
    assert interpreter.evaluate('=COUNTIF(B1:B5, "?pple")', "Sheet1") == 1
    # Combined wildcards
    assert interpreter.evaluate('=COUNTIF(B1:B5, "a*o*")', "Sheet1") == 2


# TODO: THIS TEST IS A NIGHTMARE
# def test_countif_empty_cells(interpreter):
#     """Test COUNTIF with empty cells."""
#     # Setup test data
#     sheet = interpreter.workbook.active
#     sheet["C1"] = 1
#     sheet["C2"] = None  # Empty cell
#     sheet["C3"] = ""  # Empty string
#     sheet["C4"] = 0
#     sheet["C5"] = None  # Empty cell

#     # TODO: how do we even support this thing?
#     # assert interpreter.evaluate('=COUNTIF(C1:C5, )', "Sheet1") == 1
#     assert interpreter.evaluate('=COUNTIF(C1:C5, "")', "Sheet1") == 3
#     assert interpreter.evaluate('=COUNTIF(C1:C5, "<>")', "Sheet1") == 3


def test_countif_cell_reference_criteria(interpreter):
    """Test COUNTIF with cell reference as criteria."""
    # Setup test data
    sheet = interpreter.workbook.active
    for i in range(1, 6):
        sheet[f"A{i}"] = i  # A1 to A5: 1, 2, 3, 4, 5

    sheet["D1"] = 3  # Criteria cell
    sheet["D2"] = ">2"  # Criteria cell with operator

    # Test with direct cell reference
    assert interpreter.evaluate("=COUNTIF(A1:A5, D1)", "Sheet1") == 1
    # Test with cell reference containing operator
    assert interpreter.evaluate("=COUNTIF(A1:A5, D2)", "Sheet1") == 3
    # Test by building up the pattern manually
    assert interpreter.evaluate('=COUNTIF(A1:A5, ">"&D1)', "Sheet1") == 2


def test_countif_boolean_values(interpreter):
    """Test COUNTIF with boolean values."""
    # Setup test data
    sheet = interpreter.workbook.active
    sheet["E1"] = True
    sheet["E2"] = False
    sheet["E3"] = True
    sheet["E4"] = 1
    sheet["E5"] = 0

    # Count TRUE values
    assert interpreter.evaluate("=COUNTIF(E1:E5, TRUE)", "Sheet1") == 2

    # Count FALSE values
    assert interpreter.evaluate("=COUNTIF(E1:E5, FALSE)", "Sheet1") == 1

    # Count using text representation
    assert interpreter.evaluate('=COUNTIF(E1:E5, "TRUE")', "Sheet1") == 2


@pytest.mark.skip("Could not figure out how to exactly match Excel's date formatting with openpyxl and our interpreter")
def test_countif_date_values(interpreter):
    """Test COUNTIF with date values."""
    # Setup test data
    sheet = interpreter.workbook.active
    from datetime import datetime

    sheet["F1"] = datetime(2023, 1, 1)
    sheet["F2"] = datetime(2023, 2, 15)
    sheet["F3"] = datetime(2023, 3, 10)
    sheet["F4"] = datetime(2023, 4, 5)
    sheet["F5"] = datetime(2023, 1, 20)

    # Count dates equal to a specific date
    assert interpreter.evaluate("=COUNTIF(F1:F5, F1)", "Sheet1") == 1

    # Count dates in January 2023
    assert interpreter.evaluate('=COUNTIF(F1:F5, ">DATE(2022,12,31)")', "Sheet1") == 0
    assert interpreter.evaluate('=COUNTIF(F1:F5, "<DATE(2023,2,1)")', "Sheet1") == 0

    # Count dates in a specific range
    jan_count = interpreter.evaluate(
        '=COUNTIF(F1:F5, ">DATE(2022,12,31)")-COUNTIF(F1:F5, ">=DATE(2023,2,1)")',
        "Sheet1",
    )
    assert jan_count == 0


if __name__ == "__main__":
    pytest.main([__file__])
